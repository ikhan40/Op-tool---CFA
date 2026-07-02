import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
import warnings
import io
import math
from collections import defaultdict

warnings.filterwarnings('ignore')

st.set_page_config(
    page_title="Ford CFA Flow Optimization",
    page_icon="🚗",
    layout="wide",
)

# ═══════════════════════════════════════════════════════════════════════════════
# CONFIG
# ═══════════════════════════════════════════════════════════════════════════════
# load_cfa
def load_cfa(file_bytes: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)

# load_rrr
def load_rrr(file_bytes: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)

# load_country_flow
def load_country_flow(file_bytes: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)


_SUBTOTAL_CODES           = frozenset({'FD','Total Rail','Total','Freight','Fuel'})
_COST_DEVIATION_THRESHOLD = 0.01
_MIN_VOLUME_DEFAULT       = 10
_MIN_SAVINGS              = 1.0

_RAIL_RATE_CARRIERS = [
    'NS','NS (2)','CN','CN (2)','CNA1','CNA1 (2)',
    'CSX','BN','UP','UP 2','CPA1','CP','KCS','TFM1',
    'FERR','FXE','PASH OCEAN','MATS','TOTE','AML',
    'WWL','HDGL','SCYE','MOL','GESM','EUKO','HOEG','NYKS',
]

# ═══════════════════════════════════════════════════════════════════════════════
# RAMP COORDINATES
# ═══════════════════════════════════════════════════════════════════════════════
RAMP_COORDS = {
    # ── BN/BNSF ───────────────────────────────────────────────────────────────
    'ALBUQ': {'carrier':'BN',  'name':'Albuquerque',           'city':'Albuquerque',        'st':'NM', 'country':'USA',    'lat':35.0836, 'lon':-106.6511, 'status':'Active'},
    'AMARI': {'carrier':'BN',  'name':'Amarillo',              'city':'Amarillo',           'st':'TX', 'country':'USA',    'lat':35.1934, 'lon':-101.7568, 'status':'Active'},
    'DAYTO': {'carrier':'BN',  'name':"Dayton's Bluff",        'city':'Daytons Bluff',      'st':'MN', 'country':'USA',    'lat':44.9437, 'lon': -93.0396, 'status':'Active'},
    'DIWRT': {'carrier':'BN',  'name':'Dilworth',              'city':'Dilworth',           'st':'MN', 'country':'USA',    'lat':46.8793, 'lon': -96.7078, 'status':'Active'},
    'ELMIR': {'carrier':'BN',  'name':'El Mirage',             'city':'El Mirage',          'st':'AZ', 'country':'USA',    'lat':33.6128, 'lon':-112.3265, 'status':'Active'},
    'LAURE': {'carrier':'BN',  'name':'Laurel',                'city':'Laurel',             'st':'MT', 'country':'USA',    'lat':45.6694, 'lon':-108.7706, 'status':'Active'},
    'OMAHA': {'carrier':'BN',  'name':'Omaha',                 'city':'Omaha',              'st':'NE', 'country':'USA',    'lat':41.2196, 'lon': -95.9789, 'status':'Active'},
    'ORILL': {'carrier':'BN',  'name':'Orillia',               'city':'Renton',             'st':'WA', 'country':'USA',    'lat':47.4699, 'lon':-122.2171, 'status':'Active'},
    'PORTL': {'carrier':'BN',  'name':'Portland',              'city':'Portland',           'st':'OR', 'country':'USA',    'lat':45.6224, 'lon':-122.7562, 'status':'Active'},
    'PTSAN': {'carrier':'BN',  'name':'San Diego Port',        'city':'National City',      'st':'CA', 'country':'USA',    'lat':32.6781, 'lon':-117.0992, 'status':'Active'},
    'SPOKA': {'carrier':'BN',  'name':'Spokane',               'city':'Spokane',            'st':'WA', 'country':'USA',    'lat':47.6922, 'lon':-117.3769, 'status':'Active'},
    'KCNOR': {'carrier':'BN',  'name':'Kan City BN Yard',      'city':'North Kansas City',  'st':'MO', 'country':'USA',    'lat':39.1287, 'lon': -94.5756, 'status':'Inactive'},
    'KCATS': {'carrier':'BN',  'name':'Argentine Yard',        'city':'Kansas City',        'st':'KS', 'country':'USA',    'lat':39.1141, 'lon': -94.6585, 'status':'Active'},
    'ACEQU': {'carrier':'BN',  'name':'Acequia',               'city':'Littleton',          'st':'CO', 'country':'USA',    'lat':39.5499, 'lon':-105.0780, 'status':'Active'},
    'LPCHI': {'carrier':'BN',  'name':'Logistics Park Chicago','city':'Elwood',             'st':'IL', 'country':'USA',    'lat':41.4022, 'lon': -88.1208, 'status':'Active'},
    'MPHIS': {'carrier':'BN',  'name':'Memphis',               'city':'Memphis',            'st':'TN', 'country':'USA',    'lat':35.1656, 'lon': -90.0255, 'status':'Active'},
    'MIDLO': {'carrier':'BN',  'name':'Midlothian',            'city':'Midlothian',         'st':'TX', 'country':'USA',    'lat':32.4821, 'lon': -96.9939, 'status':'Active'},
    'BIRMG': {'carrier':'BN',  'name':'Birmingham',            'city':'Birmingham',         'st':'AL', 'country':'USA',    'lat':33.5186, 'lon': -86.8104, 'status':'Active'},
    'OCKLA': {'carrier':'BN',  'name':'Oklahoma City',         'city':'Oklahoma City',      'st':'OK', 'country':'USA',    'lat':35.3395, 'lon': -97.3769, 'status':'Active'},
    'SANDI': {'carrier':'BN',  'name':'San Diego',             'city':'San Diego',          'st':'CA', 'country':'USA',    'lat':32.7157, 'lon':-117.1611, 'status':'Active'},
    'PRLND': {'carrier':'BN',  'name':'PRLND',                 'city':'Houston',            'st':'TX', 'country':'USA',    'lat':29.6647, 'lon': -95.2247, 'status':'Active'},
    # ── UP ────────────────────────────────────────────────────────────────────
    'BENIC': {'carrier':'UP',  'name':'Benicia',               'city':'Benicia',            'st':'CA', 'country':'USA',    'lat':38.0494, 'lon':-122.1586, 'status':'Active'},
    'CENTR': {'carrier':'UP',  'name':'Centreville',           'city':'Centreville',        'st':'IL', 'country':'USA',    'lat':38.5834, 'lon': -90.1232, 'status':'Inactive'},
    'CHICH': {'carrier':'UP',  'name':'Chicago Heights',       'city':'Chicago Heights',    'st':'IL', 'country':'USA',    'lat':41.5061, 'lon': -87.6355, 'status':'Active'},
    'KCMUN': {'carrier':'UP',  'name':'Muncie',                'city':'Kansas City',        'st':'KS', 'country':'USA',    'lat':39.0997, 'lon': -94.7172, 'status':'Active'},
    'LASVE': {'carrier':'UP',  'name':'Las Vegas',             'city':'Las Vegas',          'st':'NV', 'country':'USA',    'lat':36.2711, 'lon':-115.2022, 'status':'Active'},
    'MESQT': {'carrier':'UP',  'name':'Mesquite',              'city':'Dallas',             'st':'TX', 'country':'USA',    'lat':32.8402, 'lon': -96.6344, 'status':'Active'},
    'MIRLA': {'carrier':'UP',  'name':'Mira Loma',             'city':'Mira Loma',          'st':'CA', 'country':'USA',    'lat':33.9883, 'lon':-117.5236, 'status':'Active'},
    'OKLCY': {'carrier':'UP',  'name':'Oklahoma City',         'city':'Oklahoma City',      'st':'OK', 'country':'USA',    'lat':35.5022, 'lon': -97.4395, 'status':'Active'},
    'PHOEN': {'carrier':'UP',  'name':'Phoenix',               'city':'Phoenix',            'st':'AZ', 'country':'USA',    'lat':33.4195, 'lon':-112.0713, 'status':'Active'},
    'PRTAL': {'carrier':'UP',  'name':'Port Allen',            'city':'Port Allen',         'st':'LA', 'country':'USA',    'lat':30.4515, 'lon': -91.2068, 'status':'Active'},
    'REISO': {'carrier':'UP',  'name':'Reisor',                'city':'Shreveport',         'st':'LA', 'country':'USA',    'lat':32.5252, 'lon': -93.7502, 'status':'Active'},
    'ROLLA': {'carrier':'UP',  'name':'Rolla',                 'city':'Henderson',          'st':'CO', 'country':'USA',    'lat':39.9081, 'lon':-104.8624, 'status':'Active'},
    'SALIN': {'carrier':'UP',  'name':'Salinas',               'city':'Salina',             'st':'KS', 'country':'USA',    'lat':38.8403, 'lon': -97.6114, 'status':'Inactive'},
    'SALLA': {'carrier':'UP',  'name':'Salt Lake City',        'city':'Salt Lake City',     'st':'UT', 'country':'USA',    'lat':40.6965, 'lon':-111.9237, 'status':'Active'},
    'WESTF': {'carrier':'UP',  'name':'Westfield',             'city':'Houston',            'st':'TX', 'country':'USA',    'lat':29.9835, 'lon': -95.2671, 'status':'Active'},
    'PRTLN': {'carrier':'UP',  'name':'UP Portland Barn',      'city':'Portland',           'st':'OR', 'country':'USA',    'lat':45.5924, 'lon':-122.6699, 'status':'Active'},
    'MILPI': {'carrier':'UP',  'name':'Milpitas',              'city':'Milpitas',           'st':'CA', 'country':'USA',    'lat':37.4323, 'lon':-121.8996, 'status':'Active'},
    'GAVIN': {'carrier':'UP',  'name':'Gavin',                 'city':'Marion',             'st':'AR', 'country':'USA',    'lat':35.2148, 'lon': -90.1954, 'status':'Active'},
    'CBLUF': {'carrier':'UP',  'name':'Council Bluffs',        'city':'Council Bluffs',     'st':'IA', 'country':'USA',    'lat':41.2619, 'lon': -95.8608, 'status':'Active'},
    'SANRO': {'carrier':'UP',  'name':'Santa Rosa',            'city':'Santa Rosa',         'st':'NM', 'country':'USA',    'lat':34.9384, 'lon':-104.6819, 'status':'Inactive'},
    'SILBO': {'carrier':'UP',  'name':'Silver Bow',            'city':'Butte',              'st':'MT', 'country':'USA',    'lat':46.0038, 'lon':-112.5348, 'status':'Inactive'},
    'SANAN': {'carrier':'UP',  'name':'San Antonio',           'city':'San Antonio',        'st':'TX', 'country':'USA',    'lat':29.5518, 'lon': -98.4145, 'status':'Active'},
    'WCHIC': {'carrier':'UP',  'name':'West Chicago',          'city':'West Chicago',       'st':'IL', 'country':'USA',    'lat':41.8870, 'lon': -88.1992, 'status':'Active'},
    'FAIRF': {'carrier':'UP',  'name':'Fairfax',               'city':'Kansas City',        'st':'KS', 'country':'USA',    'lat':39.1141, 'lon': -94.6585, 'status':'Inactive'},
    'KENTW': {'carrier':'UP',  'name':'Kent',                  'city':'Kent',               'st':'WA', 'country':'USA',    'lat':47.3809, 'lon':-122.2348, 'status':'Inactive'},
    'SPANE': {'carrier':'UP',  'name':'Spokane',               'city':'Spokane Valley',     'st':'WA', 'country':'USA',    'lat':47.6732, 'lon':-117.2394, 'status':'Active'},
    # ── NS ────────────────────────────────────────────────────────────────────
    'ATLAN': {'carrier':'NS',  'name':'Atlanta',               'city':'Atlanta',            'st':'GA', 'country':'USA',    'lat':33.6534, 'lon': -84.4011, 'status':'Active'},
    'AVOLA': {'carrier':'NS',  'name':'Avon Lake',             'city':'Avon Lake',          'st':'OH', 'country':'USA',    'lat':41.5048, 'lon': -82.0243, 'status':'Active'},
    'AYER':  {'carrier':'NS',  'name':'Ayer',                  'city':'Ayer',               'st':'MA', 'country':'USA',    'lat':42.5612, 'lon': -71.5920, 'status':'Active'},
    'BIRMI': {'carrier':'NS',  'name':'Birmingham NS',         'city':'McCalla',            'st':'AL', 'country':'USA',    'lat':33.3901, 'lon': -87.0225, 'status':'Active'},
    'BUFFA': {'carrier':'NS',  'name':'Buffalo',               'city':'Cheektowaga',        'st':'NY', 'country':'USA',    'lat':42.8929, 'lon': -78.7548, 'status':'Active'},
    'DAVIS': {'carrier':'NS',  'name':'Davisville',            'city':'North Kingstown',    'st':'RI', 'country':'USA',    'lat':41.5789, 'lon': -71.4549, 'status':'Active'},
    'DOREM': {'carrier':'NS',  'name':'Doremus',               'city':'Newark',             'st':'NJ', 'country':'USA',    'lat':40.7357, 'lon': -74.1724, 'status':'Active'},
    'DUNDA': {'carrier':'NS',  'name':'Port of Dundalk',       'city':'Baltimore',          'st':'MD', 'country':'USA',    'lat':39.2538, 'lon': -76.5452, 'status':'Active'},
    'HEGEW': {'carrier':'NS',  'name':'Hegewisch',             'city':'Chicago',            'st':'IL', 'country':'USA',    'lat':41.6606, 'lon': -87.5557, 'status':'Active'},
    'JACKV': {'carrier':'NS',  'name':'Jacksonville NS',       'city':'Jacksonville',       'st':'FL', 'country':'USA',    'lat':30.3322, 'lon': -81.6557, 'status':'Active'},
    'KCRAN': {'carrier':'NS',  'name':'Kansas City Voltz',     'city':'Birmingham MO',      'st':'MO', 'country':'USA',    'lat':39.1469, 'lon': -94.5736, 'status':'Active'},
    'MAPLT': {'carrier':'NS',  'name':'Michigan Assembly Plant','city':'Wayne',             'st':'MI', 'country':'USA',    'lat':42.2789, 'lon': -83.3849, 'status':'Active'},
    'MECHN': {'carrier':'NS',  'name':'Mechanicville',         'city':'Mechanicville',      'st':'NY', 'country':'USA',    'lat':42.9001, 'lon': -73.6918, 'status':'Active'},
    'MELVI': {'carrier':'NS',  'name':'Melvindale',            'city':'Melvindale',         'st':'MI', 'country':'USA',    'lat':42.2789, 'lon': -83.1793, 'status':'Active'},
    'MIAMI': {'carrier':'NS',  'name':'Miami',                 'city':'Miami',              'st':'FL', 'country':'USA',    'lat':25.8703, 'lon': -80.3160, 'status':'Active'},
    'NEWAR': {'carrier':'NS',  'name':'Newark',                'city':'Newark',             'st':'NJ', 'country':'USA',    'lat':40.7420, 'lon': -74.1860, 'status':'Active'},
    'NEWOR': {'carrier':'NS',  'name':'New Orleans NS',        'city':'New Orleans',        'st':'LA', 'country':'USA',    'lat':29.9745, 'lon': -90.0782, 'status':'Active'},
    'PETER': {'carrier':'NS',  'name':'Petersburg',            'city':'Petersburg',         'st':'VA', 'country':'USA',    'lat':37.2279, 'lon': -77.4019, 'status':'Active'},
    'SHELB': {'carrier':'NS',  'name':'Shelbyville',           'city':'Shelbyville',        'st':'KY', 'country':'USA',    'lat':38.2117, 'lon': -85.2269, 'status':'Active'},
    'WENTZ': {'carrier':'NS',  'name':'Wentzville',            'city':'Wentzville',         'st':'MO', 'country':'USA',    'lat':38.8115, 'lon': -90.8529, 'status':'Active'},
    'WILMI': {'carrier':'NS',  'name':'Port Wilmington',       'city':'Wilmington',         'st':'DE', 'country':'USA',    'lat':39.7198, 'lon': -75.5596, 'status':'Active'},
    'WINSA': {'carrier':'NS',  'name':'Winston Salem',         'city':'Walkertown',         'st':'NC', 'country':'USA',    'lat':36.1715, 'lon': -80.1481, 'status':'Active'},
    'WMPHS': {'carrier':'NS',  'name':'W Memphis',             'city':'Marion',             'st':'AR', 'country':'USA',    'lat':35.2148, 'lon': -90.1954, 'status':'Active'},
    'ELKHA': {'carrier':'NS',  'name':'Elkhart',               'city':'Elkhart',            'st':'IN', 'country':'USA',    'lat':41.6820, 'lon': -85.9769, 'status':'Active'},
    'FOSTO': {'carrier':'NS',  'name':'Fostoria',              'city':'Fostoria',           'st':'OH', 'country':'USA',    'lat':41.1573, 'lon': -83.4163, 'status':'Active'},
    # ── CSX/CSXT ──────────────────────────────────────────────────────────────
    'BLOIS': {'carrier':'CSX', 'name':'Jacksonville CSX',      'city':'Jacksonville',       'st':'FL', 'country':'USA',    'lat':30.4122, 'lon': -81.6348, 'status':'Active'},
    'DIXIA': {'carrier':'CSX', 'name':'Dixiana',               'city':'Columbia',           'st':'SC', 'country':'USA',    'lat':33.7579, 'lon': -81.1076, 'status':'Active'},
    'JACVL': {'carrier':'CSX', 'name':'Jacksonville CSX 2',    'city':'Jacksonville',       'st':'FL', 'country':'USA',    'lat':30.3467, 'lon': -81.7015, 'status':'Active'},
    'JESSU': {'carrier':'CSX', 'name':'Jessup',                'city':'Jessup',             'st':'MD', 'country':'USA',    'lat':39.1454, 'lon': -76.7752, 'status':'Active'},
    'LAWRE': {'carrier':'CSX', 'name':'Lawrenceville',         'city':'Dacula',             'st':'GA', 'country':'USA',    'lat':33.9840, 'lon': -83.8988, 'status':'Active'},
    'LORTN': {'carrier':'CSX', 'name':'Lordstown',             'city':'Warren',             'st':'OH', 'country':'USA',    'lat':41.2376, 'lon': -80.8184, 'status':'Active'},
    'NASHV': {'carrier':'CSX', 'name':'Nashville',             'city':'Nashville',          'st':'TN', 'country':'USA',    'lat':36.0853, 'lon': -86.7444, 'status':'Active'},
    'NWBST': {'carrier':'CSX', 'name':'New Boston',            'city':'New Boston',         'st':'MI', 'country':'USA',    'lat':42.1553, 'lon': -83.3985, 'status':'Active'},
    'ORLAN': {'carrier':'CSX', 'name':'Orlando',               'city':'Orlando',            'st':'FL', 'country':'USA',    'lat':28.4731, 'lon': -81.3753, 'status':'Active'},
    'TAMPA': {'carrier':'CSX', 'name':'Tampa',                 'city':'Tampa',              'st':'FL', 'country':'USA',    'lat':28.0836, 'lon': -82.3948, 'status':'Active'},
    'TWIOK': {'carrier':'CSX', 'name':'Twin Oaks',             'city':'Aston',              'st':'PA', 'country':'USA',    'lat':39.8659, 'lon': -75.4355, 'status':'Active'},
    'STRAW': {'carrier':'CSX', 'name':'Strawberry Yard',       'city':'Louisville',         'st':'KY', 'country':'USA',    'lat':38.2270, 'lon': -85.6894, 'status':'Active'},
    'EASBK': {'carrier':'CSX', 'name':'East Brookfield',       'city':'East Brookfield',    'st':'MA', 'country':'USA',    'lat':42.2140, 'lon': -72.0565, 'status':'Active'},
    'ALBAN': {'carrier':'CSX', 'name':'Albany',                'city':'Selkirk',            'st':'NY', 'country':'USA',    'lat':42.5776, 'lon': -73.8318, 'status':'Active'},
    'BIRMA': {'carrier':'CSX', 'name':'Birmingham CSX',        'city':'Birmingham',         'st':'AL', 'country':'USA',    'lat':33.5721, 'lon': -86.7956, 'status':'Active'},
    'CEMDL': {'carrier':'CSX', 'name':'Cementdale',            'city':'Cincinnati',         'st':'OH', 'country':'USA',    'lat':39.2558, 'lon': -84.5767, 'status':'Active'},
    'PALMC': {'carrier':'CSX', 'name':'Palm Center',           'city':'Jupiter',            'st':'FL', 'country':'USA',    'lat':26.9342, 'lon': -80.0942, 'status':'Active'},
    'GAYLD': {'carrier':'CSX', 'name':'Gaylord',               'city':'Gaylord',            'st':'MI', 'country':'USA',    'lat':45.0275, 'lon': -84.6741, 'status':'Active'},
    # ── CN ────────────────────────────────────────────────────────────────────
    'CALGA': {'carrier':'CN',  'name':'Calgary CN',            'city':'Calgary',            'st':'AB', 'country':'Canada', 'lat':51.0447, 'lon':-114.0198, 'status':'Active'},
    'DEARB': {'carrier':'CN',  'name':'Dearborn',              'city':'Dearborn',           'st':'MI', 'country':'USA',    'lat':42.3223, 'lon': -83.1763, 'status':'Active'},
    'EDMON': {'carrier':'CN',  'name':'Edmonton CN',           'city':'Edmonton',           'st':'AB', 'country':'Canada', 'lat':53.5461, 'lon':-113.4938, 'status':'Active'},
    'FLTRK': {'carrier':'CN',  'name':'Flat Rock',             'city':'Flat Rock',          'st':'MI', 'country':'USA',    'lat':42.0997, 'lon': -83.2752, 'status':'Active'},
    'HALIF': {'carrier':'CN',  'name':'Halifax',               'city':'Eastern Passage',    'st':'NS', 'country':'Canada', 'lat':44.6867, 'lon': -63.5261, 'status':'Active'},
    'JACKO': {'carrier':'CN',  'name':'Jackson CN',            'city':'Jackson',            'st':'MS', 'country':'USA',    'lat':32.3177, 'lon': -90.1812, 'status':'Active'},
    'MARKM': {'carrier':'CN',  'name':'Markham Yard',          'city':'Harvey',             'st':'IL', 'country':'USA',    'lat':41.6100, 'lon': -87.6470, 'status':'Active'},
    'MEMPH': {'carrier':'CN',  'name':'Memphis CN',            'city':'Memphis',            'st':'TN', 'country':'USA',    'lat':35.1374, 'lon': -90.0490, 'status':'Active'},
    'MONCT': {'carrier':'CN',  'name':'Moncton',               'city':'Moncton',            'st':'NB', 'country':'Canada', 'lat':46.0878, 'lon': -64.7782, 'status':'Active'},
    'MONTR': {'carrier':'CN',  'name':'Montreal',              'city':'Saint-Laurent',      'st':'QC', 'country':'Canada', 'lat':45.5017, 'lon': -73.6752, 'status':'Active'},
    'QUEBE': {'carrier':'CN',  'name':'Quebec City',           'city':'Charny',             'st':'QC', 'country':'Canada', 'lat':46.7117, 'lon': -71.2662, 'status':'Active'},
    'SASKA': {'carrier':'CN',  'name':'Saskatoon',             'city':'Saskatoon',          'st':'SK', 'country':'Canada', 'lat':52.1332, 'lon':-106.6700, 'status':'Active'},
    'SEPTI': {'carrier':'CN',  'name':'Sept Iles',             'city':'Sept Isles',         'st':'QC', 'country':'Canada', 'lat':50.2138, 'lon': -66.3833, 'status':'Active'},
    'TORON': {'carrier':'CN',  'name':'Toronto',               'city':'Concord',            'st':'ON', 'country':'Canada', 'lat':43.7985, 'lon': -79.4944, 'status':'Active'},
    'VANCO': {'carrier':'CN',  'name':'Vancouver CN',          'city':'New Westminster',    'st':'BC', 'country':'Canada', 'lat':49.2057, 'lon':-122.9110, 'status':'Active'},
    'WINNI': {'carrier':'CN',  'name':'Winnipeg CN',           'city':'Winnipeg',           'st':'MB', 'country':'Canada', 'lat':49.8951, 'lon': -97.1384, 'status':'Active'},
    'WOODH': {'carrier':'CN',  'name':'King Road',             'city':'Brownstown Twp',     'st':'MI', 'country':'USA',    'lat':42.1267, 'lon': -83.2630, 'status':'Active'},
    'THORT': {'carrier':'CN',  'name':'Thorton Yard',          'city':'Surrey',             'st':'BC', 'country':'Canada', 'lat':49.1913, 'lon':-122.8490, 'status':'Active'},
    # ── CP ────────────────────────────────────────────────────────────────────
    'AGINC': {'carrier':'CP',  'name':'Agincourt',             'city':'Scarborough',        'st':'ON', 'country':'Canada', 'lat':43.7875, 'lon': -79.2627, 'status':'Active'},
    'REGIN': {'carrier':'CP',  'name':'Regina',                'city':'Regina',             'st':'SK', 'country':'Canada', 'lat':50.4452, 'lon':-104.6189, 'status':'Active'},
    'RGEBA': {'carrier':'CP',  'name':'Richards Gebaur',       'city':'Kansas City',        'st':'MO', 'country':'USA',    'lat':38.8814, 'lon': -94.5133, 'status':'Active'},
    'COTGR': {'carrier':'CP',  'name':'Cottage Grove',         'city':'Cottage Grove',      'st':'MN', 'country':'USA',    'lat':44.8319, 'lon': -92.9385, 'status':'Active'},
    'VCOUV': {'carrier':'CP',  'name':'Vancouver CP',          'city':'Pitt Meadows',       'st':'BC', 'country':'Canada', 'lat':49.2190, 'lon':-122.6890, 'status':'Active'},
    'SCHLR': {'carrier':'CP',  'name':'Schiller Park',         'city':'Schiller Park',      'st':'IL', 'country':'USA',    'lat':41.9559, 'lon': -87.8756, 'status':'Active'},
    'WINDS': {'carrier':'CP',  'name':'Windsor',               'city':'Windsor',            'st':'ON', 'country':'Canada', 'lat':42.3149, 'lon': -83.0364, 'status':'Active'},
    'WYLIE': {'carrier':'CP',  'name':'Wylie',                 'city':'Wylie',              'st':'TX', 'country':'USA',    'lat':33.0151, 'lon': -96.5388, 'status':'Active'},
    'ELLER': {'carrier':'CP',  'name':'Ellerslie',             'city':'Edmonton',           'st':'AB', 'country':'Canada', 'lat':53.4556, 'lon':-113.5559, 'status':'Active'},
    'MONRE': {'carrier':'CP',  'name':'Monterrey CP',          'city':'Monterrey',          'st':'NL', 'country':'Mexico', 'lat':25.6866, 'lon':-100.3161, 'status':'Active'},
    # ── FXE / KCS ─────────────────────────────────────────────────────────────
    'GUADA': {'carrier':'FXE', 'name':'Guadalajara',           'city':'Guadalajara',        'st':'JAL','country':'Mexico', 'lat':20.6597, 'lon':-103.3496, 'status':'Active'},
    'HERSO': {'carrier':'FXE', 'name':'Hermosillo Inbound',    'city':'Hermosillo',         'st':'SO', 'country':'Mexico', 'lat':29.0730, 'lon':-110.9559, 'status':'Active'},
    'HERMO': {'carrier':'FXE', 'name':'Hermosillo',            'city':'Hermosillo',         'st':'SO', 'country':'Mexico', 'lat':29.0850, 'lon':-110.9400, 'status':'Active'},
    'MEXIC': {'carrier':'FXE', 'name':'Monterrey FXE',         'city':'Monterrey',          'st':'NL', 'country':'Mexico', 'lat':25.7200, 'lon':-100.2800, 'status':'Active'},
    'SANRI': {'carrier':'FXE', 'name':'Santa Rita',            'city':'Santa Rita',         'st':'VE', 'country':'Mexico', 'lat':19.1760, 'lon': -96.2522, 'status':'Active'},
    'KENDL': {'carrier':'KCS', 'name':'Kendleton',             'city':'Kendleton',          'st':'TX', 'country':'USA',    'lat':29.4488, 'lon': -95.9611, 'status':'Active'},
    'ENVIO': {'carrier':'KCS', 'name':'Envios Cuautitlan',     'city':'Cuautitlan Izcalli', 'st':'MEX','country':'Mexico', 'lat':19.6630, 'lon': -99.2097, 'status':'Active'},
    'ADESA': {'carrier':'None','name':'Adesa Phoenix',         'city':'Chandler',           'st':'AZ', 'country':'USA',    'lat':33.3062, 'lon':-111.8413, 'status':'Active'},
}


def haversine(lat1, lon1, lat2, lon2):
    """Great-circle distance in miles."""
    R    = 3958.8
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    a    = (math.sin((phi2 - phi1) / 2) ** 2 +
            math.cos(phi1) * math.cos(phi2) *
            math.sin(math.radians(lon2 - lon1) / 2) ** 2)
    return 2 * R * math.asin(math.sqrt(a))


# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════
def clean_num(v):
    if v is None: return 0.0
    if isinstance(v, float) and np.isnan(v): return 0.0
    if isinstance(v, (int, float, np.integer, np.floating)): return float(v)
    s = str(v).strip()
    if s in ('','#N/A','nan','None','N/A','#VALUE!','#REF!',
             '#DIV/0!','-','—',' -   ',' - ','$-   ','$-'): return 0.0
    if s.startswith('='): return 0.0
    s = s.replace('$','').replace(',','').replace(' ','')
    if s.startswith('(') and s.endswith(')'): s = '-' + s[1:-1]
    try: return float(s)
    except: return 0.0

def fmt_c(v):
    try:
        f = float(v)
        if f == 0: return '$—'
        return f"${f:,.0f}" if f >= 0 else f"(${abs(f):,.0f})"
    except: return '—'

def fmt_r(v, decimals=2):
    try:
        f = float(v)
        if f == 0: return '—'
        return f"${f:,.{decimals}f}"
    except: return '—'

def safe_str(v):
    s = str(v or '').strip()
    return '' if s.lower() in ('nan','none') else s

def normalize_col(v):
    s = str(v).strip()
    try: return str(int(float(s)))
    except: return s

def normalize_fips(v):
    try: return str(int(float(str(v).strip().replace(',','')))).zfill(5)
    except: return '00000'

def find_col(df, name):
    return next((c for c in df.columns if c.strip().lower() == name.lower()), None)

def label_flow(flow):
    s = str(flow or '').strip()
    if s.upper() in ('','N/A','NA','NONE','0','DEALER DIRECT','DD'):
        return 'Dealer Direct'
    return s

def flow_display(flow, ramp, other_flow, other_ramp):
    base       = label_flow(flow)
    other_base = label_flow(other_flow)
    if base == 'Dealer Direct' and other_base == 'Dealer Direct':
        return f"Dealer Direct  [{ramp}]"
    if base != other_base:
        return base
    return f"{base}  [{ramp}]"

def describe_change(cur_cr, pro_cr):
    changes = []
    cur_fd = (cur_cr.get('fd_carrier') or '').strip()
    pro_fd = (pro_cr.get('fd_carrier') or '').strip()

    if cur_fd != pro_fd:
        changes.append(f"FD: {cur_fd or '—'} → {pro_fd or '—'}")

    cur_rail = set(k for k,v in cur_cr.get('rail_breakdown',{}).items()
                   if clean_num(v.get('freight',0)) > 0)
    pro_rail = set(k for k,v in pro_cr.get('rail_breakdown',{}).items()
                   if clean_num(v.get('freight',0)) > 0)

    fd_carriers = {cur_fd, pro_fd} - {''}
    lost   = (cur_rail - pro_rail) - fd_carriers
    gained = (pro_rail - cur_rail) - fd_carriers

    if lost:   changes.append(f"Drop: {', '.join(sorted(lost))}")
    if gained: changes.append(f"Add: {', '.join(sorted(gained))}")

    cur_sh = set(k for k,v in cur_cr.get('sh_breakdown',{}).items()
                 if clean_num(v.get('freight',0)) > 0)
    pro_sh = set(k for k,v in pro_cr.get('sh_breakdown',{}).items()
                 if clean_num(v.get('freight',0)) > 0)
    sh_lost   = cur_sh - pro_sh
    sh_gained = pro_sh - cur_sh
    if sh_lost:   changes.append(f"Drop Shuttle: {', '.join(sorted(sh_lost))}")
    if sh_gained: changes.append(f"Add Shuttle: {', '.join(sorted(sh_gained))}")

    return ' | '.join(changes) if changes else 'Rate / volume adjustment only'


# ═══════════════════════════════════════════════════════════════════════════════
# PER-LEG DETAIL
# ═══════════════════════════════════════════════════════════════════════════════
def get_rail_legs(row_dict, scale=1.0):
    legs = []
    for carrier in _RAIL_RATE_CARRIERS:
        rate    = clean_num(row_dict.get(f"{carrier} Rate",    0))
        miles   = clean_num(row_dict.get(f"{carrier} Miles",   0))
        freight = clean_num(row_dict.get(f"{carrier} Freight", 0))
        fuel    = clean_num(row_dict.get(f"{carrier} Fuel",    0))
        if freight > 0 or rate > 0:
            legs.append({
                'carrier' : carrier,
                'rate_pu' : rate,
                'miles'   : int(round(miles)),
                'freight' : freight * scale,
                'fuel'    : fuel    * scale,
                'total'   : (freight + fuel) * scale,
            })
    legs.sort(key=lambda x: -x['freight'])
    return legs


def get_shuttle_legs(row_dict, scale=1.0):
    legs = []
    for i in (1, 2):
        carrier = safe_str(row_dict.get(f'Shuttle Carrier {i}', ''))
        miles   = clean_num(row_dict.get(f'Shuttle Miles {i}', 0))
        flow    = safe_str(row_dict.get(f'Shuttle Flow {i}', ''))
        if not carrier: continue
        freight = clean_num(row_dict.get(f'{carrier} Shuttle Freight', 0))
        fuel    = clean_num(row_dict.get(f'{carrier} Shuttle Fuel',    0))
        legs.append({
            'leg'     : i,
            'carrier' : carrier,
            'flow'    : flow,
            'miles'   : int(round(miles)),
            'freight' : freight * scale,
            'fuel'    : fuel    * scale,
            'total'   : (freight + fuel) * scale,
        })
    return legs


def get_fd_detail(row_dict, scale=1.0):
    freight = clean_num(row_dict.get('FD Freight', 0))
    fuel    = clean_num(row_dict.get('FD Fuel',    0))
    return {
        'carrier'  : safe_str(row_dict.get('FD Carrier', '')),
        'fixed_pu' : clean_num(row_dict.get('FD Fixed', 0)),
        'var_pu'   : clean_num(row_dict.get('FD Var',   0)),
        'miles'    : clean_num(row_dict.get('Miles',    0)),
        'freight'  : freight * scale,
        'fuel'     : fuel    * scale,
    }


def get_identity(row_dict, scale=1.0):
    return {
        'fips'             : safe_str(row_dict.get('Fips',    '')),
        'county'           : safe_str(row_dict.get('County',  '')).title(),
        'st'               : safe_str(row_dict.get('ST',      '')),
        'rep_city'         : safe_str(row_dict.get('Rep City','')).title(),
        'rc'               : safe_str(row_dict.get('RC',      '')),
        'pt'               : safe_str(row_dict.get('P/T',     '')),
        'rrc'              : safe_str(row_dict.get('RRC',     '')),
        'ramp'             : safe_str(row_dict.get('Ramp',    '')).upper(),
        'flow'             : safe_str(row_dict.get('Flow',    '')),
        'fd_carrier'       : safe_str(row_dict.get('FD Carrier','')),
        'volume'           : clean_num(row_dict.get('Volume', 0)),
        'miles'            : clean_num(row_dict.get('Miles',  0)),
        'railcars'         : clean_num(row_dict.get('Railcars', 0) or
                                       row_dict.get(' Railcars ', 0)),
        'carhaul_rigs'     : clean_num(row_dict.get('Carhaul Rigs', 0) or
                                       row_dict.get(' Carhaul Rigs ', 0)),
        'clr'              : clean_num(row_dict.get('CLR', 0) or
                                       row_dict.get(' CLR ', 0)),
        'rlr'              : clean_num(row_dict.get('RLR', 0) or
                                       row_dict.get(' RLR ', 0)),
        'origin_rr'        : safe_str(row_dict.get('Origin Railroad','')),
        'dest_rr'          : safe_str(row_dict.get('Destination Railroad','')),
        'o_lat'            : clean_num(row_dict.get('O Latitude',  0)),
        'o_lon'            : clean_num(row_dict.get('O Longitude', 0)),
        'd_lat'            : clean_num(row_dict.get('D Latitude',  0) or
                                       row_dict.get('Latitude',    0)),
        'd_lon'            : clean_num(row_dict.get('D Longitude', 0) or
                                       row_dict.get('Longitude',   0)),
        'total_rail_miles' : clean_num(row_dict.get('Total Rail Miles per unit', 0)),
        'total_miles'      : clean_num(row_dict.get('Total Miles', 0)),
        'cpu'              : clean_num(row_dict.get('CPU', 0)),
        'ramp_freight'     : clean_num(row_dict.get('Ramp Freight', 0) or
                                       row_dict.get(' Ramp Freight ', 0)) * scale,
        'col_freight'      : (clean_num(row_dict.get(' Freight ', 0)) or
                              clean_num(row_dict.get('Freight',   0))) * scale,
        'col_fuel'         : (clean_num(row_dict.get(' Fuel ', 0)) or
                              clean_num(row_dict.get('Fuel',   0))) * scale,
        'col_total'        : (clean_num(row_dict.get(' Total ', 0)) or
                              clean_num(row_dict.get('Total',   0))) * scale,
    }


def _compute_scale(row_dict, target_volume):
    cfa_vol = clean_num(row_dict.get('Volume', 0))
    if cfa_vol > 0.000001:
        return target_volume / cfa_vol
    return target_volume


# ═══════════════════════════════════════════════════════════════════════════════
# MAP
# ═══════════════════════════════════════════════════════════════════════════════
def render_map(o_lat, o_lon, d_lat, d_lon, origin_label, dest_label):
    try:
        if not all([o_lat, o_lon, d_lat, d_lon]): return
        o_lat, o_lon = float(o_lat), float(o_lon)
        d_lat, d_lon = float(d_lat), float(d_lon)
        if abs(o_lat) < 0.1 or abs(d_lat) < 0.1: return

        import pydeck as pdk

        pts = pd.DataFrame([
            {'lat': o_lat, 'lon': o_lon, 'name': f"🏭 {origin_label}", 'r':0,   'g':100,'b':255},
            {'lat': d_lat, 'lon': d_lon, 'name': f"📍 {dest_label}",   'r':255, 'g':50, 'b':50},
        ])
        arcs = pd.DataFrame([{
            'slat': o_lat, 'slon': o_lon,
            'tlat': d_lat, 'tlon': d_lon,
        }])
        scatter = pdk.Layer('ScatterplotLayer', data=pts,
                            get_position=['lon','lat'],
                            get_fill_color=['r','g','b'],
                            get_radius=20000, pickable=True)
        arc     = pdk.Layer('ArcLayer', data=arcs,
                            get_source_position=['slon','slat'],
                            get_target_position=['tlon','tlat'],
                            get_source_color=[0,100,255,180],
                            get_target_color=[255,50,50,180],
                            get_width=4)
        dist = ((o_lat-d_lat)**2 + (o_lon-d_lon)**2)**0.5
        zoom = max(3, min(8, 8 - dist * 0.4))
        st.pydeck_chart(pdk.Deck(
            layers=[arc, scatter],
            initial_view_state=pdk.ViewState(
                latitude=(o_lat+d_lat)/2,
                longitude=(o_lon+d_lon)/2,
                zoom=zoom, pitch=20,
            ),
            tooltip={'text': '{name}'},
        ))
    except Exception as e:
        try:
            map_df = pd.DataFrame([
                {'lat': float(o_lat), 'lon': float(o_lon)},
                {'lat': float(d_lat), 'lon': float(d_lon)},
            ])
            st.map(map_df, zoom=4)
        except:
            st.caption(f"Map unavailable ({e})")


# ═══════════════════════════════════════════════════════════════════════════════
# LOADERS
# ═══════════════════════════════════════════════════════════════════════════════
def _read_sheet(ws):
    rows = list(ws.values)
    if len(rows) < 2: return pd.DataFrame()
    seen, hdrs = {}, []
    for i, h in enumerate(rows[0]):
        n   = str(h).strip() if h else f'_C{i}'
        cnt = seen.get(n, 0); seen[n] = cnt + 1
        hdrs.append(n if cnt == 0 else f"{n}_{cnt}")
    return pd.DataFrame(rows[1:], columns=hdrs).dropna(how='all')


@st.cache_data
def load_cfa(file_bytes: bytes):
    wb  = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    df  = _read_sheet(wb.active); wb.close()
    str_cols = ['Origin','Alt','Ramp','Flow','FD Carrier','County','ST',
                'Column','P/T','Shuttle Carrier 1','Shuttle Carrier 2',
                'Shuttle Flow 1','Shuttle Flow 2']
    for col in str_cols:
        if col in df.columns:
            df[col] = df[col].fillna('').astype(str).str.strip()
    df['Fips'] = (pd.to_numeric(df['Fips'], errors='coerce')
                  .fillna(0).astype(int).astype(str).str.zfill(5))
    for c in ('Flow','Alt'):
        if c in df.columns:
            df[c] = df[c].replace({'nan':'','None':'','none':''})
    df['_vol']      = df['Volume'].apply(clean_num)
    df['_col_norm'] = df['Column'].apply(normalize_col)
    df = df[df['Fips'] != '00000'].copy().reset_index(drop=True)
    return df


@st.cache_data
def load_rrr(file_bytes: bytes):
    wb      = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    buckets = {'Ramp':[],'Terr':[],'City':[],'Via':[],'Ramp Charges':[]}
    for sheet_name in wb.sheetnames:
        df = _read_sheet(wb[sheet_name])
        if df.empty: continue
        screen_col = find_col(df, 'Screen')
        if screen_col:
            for sv, sub in df.groupby(screen_col):
                sv = str(sv).strip()
                if sv in buckets: buckets[sv].append(sub)
        else:
            hdrs = set(c.strip() for c in df.columns)
            if   'Ramp Mnem' in hdrs and 'Charge Type' in hdrs: buckets['Ramp Charges'].append(df)
            elif 'Orig Ramp Mnem' in hdrs:                       buckets['Via'].append(df)
            elif any('FIPS' in c.upper() for c in hdrs):         buckets['Terr'].append(df)
            elif any('SPLC' in c.upper() for c in hdrs):         buckets['City'].append(df)
            else:                                                 buckets['Ramp'].append(df)
    wb.close()
    def concat(lst): return pd.concat(lst, ignore_index=True) if lst else pd.DataFrame()
    r2r=concat(buckets['Ramp']); r2t=concat(buckets['Terr'])
    r2c=concat(buckets['City']); via=concat(buckets['Via'])
    rc =concat(buckets['Ramp Charges'])
    if not r2r.empty:
        oc=find_col(r2r,'Orig Ramp'); dc=find_col(r2r,'Dest Ramp')
        cc=find_col(r2r,'Rate Col');  af=find_col(r2r,'Active Flag Route')
        if oc: r2r['_orig_ramp']=r2r[oc].fillna('').astype(str).str.strip().str.upper()
        if dc: r2r['_dest_ramp']=r2r[dc].fillna('').astype(str).str.strip().str.upper()
        if cc: r2r['_rate_col'] =r2r[cc].apply(normalize_col)
        if af: r2r=r2r[r2r[af].apply(clean_num)==0].copy()
    if not r2t.empty:
        oc=find_col(r2t,'Orig Ramp'); fc=find_col(r2t,'Dest FIPS')
        cc=find_col(r2t,'Rate Col');  af=find_col(r2t,'Active Flag Route')
        if oc: r2t['_orig_ramp']=r2t[oc].fillna('').astype(str).str.strip().str.upper()
        if fc: r2t['_fips']     =r2t[fc].apply(normalize_fips)
        if cc: r2t['_rate_col'] =r2t[cc].apply(normalize_col)
        if af: r2t=r2t[r2t[af].apply(clean_num)==0].copy()
    if not rc.empty:
        rm=find_col(rc,'Ramp Mnem'); rt=find_col(rc,'Rate'); af=find_col(rc,'Active Flag')
        if rm: rc['_ramp']=rc[rm].fillna('').astype(str).str.strip().str.upper()
        if rt: rc['_rate']=rc[rt].apply(clean_num)
        if af: rc=rc[rc[af].apply(clean_num)==0].copy()
    return {'r2r':r2r,'r2t':r2t,'r2c':r2c,'via':via,'rc':rc}


@st.cache_data
def load_country_flow(file_bytes: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    df = _read_sheet(wb.active); wb.close()
    str_cols = ['Origin','RC','Ramp','Flow','FD Carrier','County','ST','Column','P/T','Region']
    for col in str_cols:
        if col in df.columns:
            df[col] = df[col].fillna('').astype(str).str.strip()
    df['Fips']      = df['Fips'].apply(normalize_fips)
    df['_vol']      = df['Volume'].apply(clean_num)
    df['_miles']    = df['Miles'].apply(clean_num)
    df['_col_norm'] = df['Column'].apply(normalize_col)
    df = df[df['_vol']  > 0].copy()
    df = df[df['Fips'] != '00000'].copy().reset_index(drop=True)
    return df


# ═══════════════════════════════════════════════════════════════════════════════
# COST CALCULATOR
# ═══════════════════════════════════════════════════════════════════════════════
def get_ramp_col_pairs(df_cols):
    pairs = []
    for col in df_cols:
        cs = col.strip()
        if not cs.endswith(' Freight') or 'Shuttle' in cs: continue
        code = cs[:-8].strip()
        if not code or code in _SUBTOTAL_CODES: continue
        if 'total' in code.lower(): continue
        fuel_col = next(
            (c for c in df_cols if c.strip() == f"{code} Fuel"),
            col.replace('Freight','Fuel'),
        )
        pairs.append((code, col, fuel_col))
    return pairs


def get_shuttle_col_pairs(df_cols):
    pairs = []
    for col in df_cols:
        cs = col.strip()
        if not cs.endswith(' Shuttle Freight'): continue
        code = cs[:-16].strip()
        if not code: continue
        fuel_col = next(
            (c for c in df_cols if c.strip() == f"{code} Shuttle Fuel"),
            col.replace('Shuttle Freight','Shuttle Fuel'),
        )
        pairs.append((code, col, fuel_col))
    return pairs


def get_all_carriers(row_dict, ramp_pairs, shuttle_pairs):
    carriers = []
    for code, fc, fuc in ramp_pairs:
        if clean_num(row_dict.get(fc,0)) > 0 or clean_num(row_dict.get(fuc,0)) > 0:
            carriers.append(code)
    for code, fc, fuc in shuttle_pairs:
        if clean_num(row_dict.get(fc,0)) > 0 or clean_num(row_dict.get(fuc,0)) > 0:
            carriers.append(code)
    return list(dict.fromkeys(carriers))


def calculate_cfa_cost(row_dict, target_volume, ramp_pairs, shuttle_pairs):
    row_vol       = clean_num(row_dict.get('Volume', 0))
    target_volume = max(target_volume, 0.000001)
    scale         = (target_volume / row_vol) if row_vol > 0.000001 else target_volume

    total_freight = 0.0; total_fuel = 0.0
    rail_breakdown = {}; sh_breakdown = {}; has_shuttle = False

    for code, fc, fuc in ramp_pairs:
        freight = clean_num(row_dict.get(fc,0))
        fuel    = clean_num(row_dict.get(fuc,0))
        if freight > 0 or fuel > 0:
            sf = freight * scale; uf = fuel * scale
            total_freight += sf; total_fuel += uf
            rail_breakdown[code] = {'freight': sf, 'fuel': uf}

    for code, fc, fuc in shuttle_pairs:
        freight = clean_num(row_dict.get(fc,0))
        fuel    = clean_num(row_dict.get(fuc,0))
        if freight > 0 or fuel > 0:
            has_shuttle = True
            sf = freight * scale; uf = fuel * scale
            total_freight += sf; total_fuel += uf
            sh_breakdown[code] = {'freight': sf, 'fuel': uf}

    our_total = total_freight + total_fuel

    cfa_freight = (clean_num(row_dict.get(' Freight ',0)) or
                   clean_num(row_dict.get('Freight',0)))
    cfa_fuel    = (clean_num(row_dict.get(' Fuel ',0)) or
                   clean_num(row_dict.get('Fuel',0)))
    cfa_raw     = cfa_freight + cfa_fuel
    cost_adjusted = False

    if cfa_raw > 0 and row_vol > 0.000001:
        cfa_scaled = cfa_raw * scale
        deviation  = abs(our_total - cfa_scaled) / max(cfa_scaled, 1)
        if deviation > _COST_DEVIATION_THRESHOLD:
            adj = cfa_scaled / max(our_total, 0.000001)
            for code in rail_breakdown:
                rail_breakdown[code]['freight'] *= adj
                rail_breakdown[code]['fuel']    *= adj
            for code in sh_breakdown:
                sh_breakdown[code]['freight'] *= adj
                sh_breakdown[code]['fuel']    *= adj
            total_freight *= adj; total_fuel *= adj
            our_total = cfa_scaled; cost_adjusted = True

    return {
        'total'          : our_total,
        'total_freight'  : total_freight,
        'total_fuel'     : total_fuel,
        'rail_breakdown' : rail_breakdown,
        'sh_breakdown'   : sh_breakdown,
        'has_shuttle'    : has_shuttle,
        'sh_carrier1'    : safe_str(row_dict.get('Shuttle Carrier 1','')),
        'sh_carrier2'    : safe_str(row_dict.get('Shuttle Carrier 2','')),
        'sh_miles1'      : clean_num(row_dict.get('Shuttle Miles 1',0)),
        'sh_miles2'      : clean_num(row_dict.get('Shuttle Miles 2',0)),
        'fd_carrier'     : safe_str(row_dict.get('FD Carrier','')),
        'fd_miles'       : clean_num(row_dict.get('Miles',0)),
        'source'         : 'CFA',
        'cfa_volume'     : row_vol,
        'volume_adjusted': abs(row_vol - target_volume) > 0.01,
        'cost_adjusted'  : cost_adjusted,
    }


def get_ramp_charges_total(dest_ramp, rc_df):
    if rc_df.empty or '_ramp' not in rc_df.columns: return 0.0
    mask = rc_df['_ramp'] == dest_ramp.upper().strip()
    return float(rc_df.loc[mask, '_rate'].sum())


# ═══════════════════════════════════════════════════════════════════════════════
# COMPARISON BUILDER
# ═══════════════════════════════════════════════════════════════════════════════
def _pick_best_cfa_row(cfa_matches, cf_flow):
    if not cfa_matches: return None
    cf_flow_clean = cf_flow.strip().upper()
    for row in cfa_matches:
        if safe_str(row.get('Flow','')).upper() == cf_flow_clean: return row
    return max(cfa_matches, key=lambda r: clean_num(r.get('Volume',0)))


def _type_label(r):
    cur_dd = label_flow(r['current_flow'])  == 'Dealer Direct'
    pro_dd = label_flow(r['proposed_flow']) == 'Dealer Direct'
    if   cur_dd and not pro_dd: return '🚂 DD→Rail'
    elif not cur_dd and pro_dd: return '🚗 Rail→DD'
    else:                       return '🔄 Rail→Rail'


def _make_entry(orig, fips, county, st_code, col, rc, cf_vol,
                cur_flow, cur_ramp, cur_cr, cur_total,
                pro_flow, pro_ramp, pro_cr, pro_total,
                savings, same_ramp_flag,
                ramp_pairs, shuttle_pairs, source,
                cur_raw=None, pro_raw=None, flags=None):
    sav_pct   = (savings / cur_total * 100) if cur_total > 0 else 0.0
    vol       = max(cf_vol, 1)
    what_chgd = describe_change(cur_cr, pro_cr) if same_ramp_flag else ''
    return {
        'origin'              : orig,
        'fips'                : fips,
        'county'              : county,
        'st'                  : st_code,
        'column'              : col,
        'rc'                  : rc,
        'volume'              : cf_vol,
        'current_flow'        : cur_flow,
        'current_ramp'        : cur_ramp,
        'current_cost'        : round(cur_total, 2),
        'current_cpu'         : round(cur_total / vol, 2),
        'current_fd_carrier'  : (cur_cr or {}).get('fd_carrier',''),
        'current_breakdown'   : cur_cr or {},
        'current_raw'         : cur_raw or {},
        'proposed_flow'       : pro_flow,
        'proposed_ramp'       : pro_ramp,
        'proposed_cost'       : round(pro_total, 2),
        'proposed_cpu'        : round(pro_total / vol, 2),
        'proposed_fd_carrier' : (pro_cr or {}).get('fd_carrier',''),
        'proposed_breakdown'  : pro_cr or {},
        'proposed_raw'        : pro_raw or {},
        'what_changed'        : what_chgd,
        'savings'             : round(savings, 2),
        'savings_pct'         : round(sav_pct, 1),
        'cpu_change'          : round((pro_total - cur_total) / vol, 2),
        'same_ramp'           : same_ramp_flag,
        'source'              : source,
        'flags'               : flags or [],
        'ramp_pairs'          : ramp_pairs,
        'shuttle_pairs'       : shuttle_pairs,
    }


@st.cache_data
def build_comparisons(_cf_df, _cfa_df, _r2r_df, _r2t_df, _rc_df):
    df_cols       = _cfa_df.columns.tolist()
    ramp_pairs    = get_ramp_col_pairs(df_cols)
    shuttle_pairs = get_shuttle_col_pairs(df_cols)

    cfa_index = defaultdict(list)
    for _, row in _cfa_df.iterrows():
        key = (row['Origin'].upper(), row['Fips'],
               normalize_col(row['Column']), row['Ramp'].upper())
        cfa_index[key].append(row.to_dict())

    r2r_index = {}
    if not _r2r_df.empty and '_orig_ramp' in _r2r_df.columns:
        for _, row in _r2r_df.iterrows():
            key  = (row.get('_orig_ramp',''), row.get('_dest_ramp',''), row.get('_rate_col',''))
            if not key[0] or not key[1]: continue
            rate = clean_num(row.get('Fixed Rate', float('inf')))
            if key not in r2r_index or rate < clean_num(r2r_index[key].get('Fixed Rate', float('inf'))):
                r2r_index[key] = row.to_dict()

    r2t_index          = {}
    fips_to_dest_ramps = defaultdict(set)
    if not _r2t_df.empty and '_orig_ramp' in _r2t_df.columns:
        for _, row in _r2t_df.iterrows():
            dr=row.get('_orig_ramp',''); fips=row.get('_fips',''); col=row.get('_rate_col','')
            if not dr or not fips: continue
            key  = (dr, fips, col)
            rate = clean_num(row.get('Fixed Rate', float('inf')))
            if key not in r2t_index or rate < clean_num(r2t_index[key].get('Fixed Rate', float('inf'))):
                r2t_index[key] = row.to_dict()
            fips_to_dest_ramps[fips].add(dr)

    _rc_cache = {}
    def get_rc(ramp):
        if ramp not in _rc_cache:
            _rc_cache[ramp] = get_ramp_charges_total(ramp, _rc_df)
        return _rc_cache[ramp]

    carrier_vol_total = defaultdict(float)
    carrier_vol_ramp  = defaultdict(float)
    for _, cfrow in _cf_df.iterrows():
        orig=cfrow['Origin'].upper(); fips=cfrow['Fips']
        col=normalize_col(cfrow['Column']); ramp=cfrow['Ramp'].upper(); vol=cfrow['_vol']
        for match in cfa_index.get((orig,fips,col,ramp),[]):
            for c in get_all_carriers(match, ramp_pairs, shuttle_pairs):
                carrier_vol_total[c]        += vol
                carrier_vol_ramp[(c, ramp)] += vol

    same_ramp_results = []; diff_ramp_results = []

    for _, cfrow in _cf_df.iterrows():
        orig     = cfrow['Origin'].upper()
        fips     = cfrow['Fips']
        col      = normalize_col(cfrow['Column'])
        ramp     = cfrow['Ramp'].upper()
        cf_vol   = cfrow['_vol']
        cf_miles = cfrow['_miles']
        county   = str(cfrow.get('County','')).strip().title()
        st_code  = str(cfrow.get('ST','')).strip()
        cf_flow  = safe_str(cfrow.get('Flow',''))
        rc_code  = safe_str(cfrow.get('RC',''))
        if cf_vol <= 0: continue

        cfa_key     = (orig, fips, col, ramp)
        cfa_matches = cfa_index.get(cfa_key, [])
        if not cfa_matches:
            for k, v in cfa_index.items():
                if k[0]==orig and k[1]==fips and k[2]==col:
                    cfa_matches=v; break
        if not cfa_matches: continue

        cur_row = _pick_best_cfa_row(cfa_matches, cf_flow)
        if cur_row is None: continue

        cur_cr    = calculate_cfa_cost(cur_row, cf_vol, ramp_pairs, shuttle_pairs)
        cur_total = cur_cr['total']
        cur_flow  = safe_str(cur_row.get('Flow','')) or cf_flow
        if cur_total <= 0: continue

        same_alts = []; diff_alts = []

        for k, rows in cfa_index.items():
            if k[0]!=orig or k[1]!=fips or k[2]!=col: continue
            alt_ramp       = k[3]
            same_ramp_flag = (alt_ramp == ramp)
            for alt_row in rows:
                if (same_ramp_flag and
                        alt_row.get('Flow','')  == cur_row.get('Flow','') and
                        clean_num(alt_row.get('Volume',0)) == clean_num(cur_row.get('Volume',0))):
                    continue
                alt_cr  = calculate_cfa_cost(alt_row, cf_vol, ramp_pairs, shuttle_pairs)
                alt_tot = alt_cr['total']
                if alt_tot <= 0: continue
                savings = cur_total - alt_tot
                if savings < _MIN_SAVINGS: continue
                alt_flow    = safe_str(alt_row.get('Flow',''))
                entry_flags = []
                if alt_cr.get('cost_adjusted'):
                    entry_flags.append("Proposed cost adjusted to CFA summary — verify carrier breakdown.")
                entry = _make_entry(
                    orig, fips, county, st_code, col, rc_code, cf_vol,
                    cur_flow, ramp, cur_cr, cur_total,
                    alt_flow, alt_ramp, alt_cr, alt_tot,
                    savings, same_ramp_flag, ramp_pairs, shuttle_pairs, 'CFA',
                    cur_raw=cur_row, pro_raw=alt_row, flags=entry_flags,
                )
                (same_alts if same_ramp_flag else diff_alts).append(entry)

        for dest_ramp in fips_to_dest_ramps.get(fips, set()):
            r2r_key=(orig, dest_ramp, col); r2t_key=(dest_ramp, fips, col)
            if r2r_key not in r2r_index or r2t_key not in r2t_index: continue
            r2r_row=r2r_index[r2r_key]; r2t_row=r2t_index[r2t_key]
            rail_rate =clean_num(r2r_row.get('Fixed Rate',0))
            fd_fixed  =clean_num(r2t_row.get('Fixed Rate',0))
            fd_var    =clean_num(r2t_row.get('Var Rate',  0))
            ramp_ch_pu=get_rc(dest_ramp)
            if rail_rate <= 0: continue
            cost_pu=rail_rate + fd_fixed + fd_var*cf_miles + ramp_ch_pu
            alt_tot=cost_pu*cf_vol; savings=cur_total-alt_tot
            if savings < _MIN_SAVINGS: continue
            rail_carrier=safe_str(r2r_row.get('PayTo Scac','')); fd_carrier=safe_str(r2t_row.get('Svc SCAC',''))
            pro_cr={
                'total':alt_tot,'total_freight':alt_tot,'total_fuel':0.0,
                'rail_breakdown':{rail_carrier:{'freight':rail_rate*cf_vol,'fuel':0.0}},
                'sh_breakdown':{},'has_shuttle':False,
                'sh_carrier1':'','sh_carrier2':'','sh_miles1':0.0,'sh_miles2':0.0,
                'fd_carrier':fd_carrier,'fd_miles':cf_miles,
                'fd_fixed_pu':fd_fixed,'fd_var_pu':fd_var,'ramp_charges_pu':ramp_ch_pu,
                'rail_carrier':rail_carrier,'source':'RRR',
                'volume_adjusted':False,'cost_adjusted':False,
            }
            pro_flow=f"{orig}  {rail_carrier}  {dest_ramp}  {fd_carrier}"
            same_ramp_flag=(dest_ramp==ramp)
            entry=_make_entry(
                orig, fips, county, st_code, col, rc_code, cf_vol,
                cur_flow, ramp, cur_cr, cur_total,
                pro_flow, dest_ramp, pro_cr, alt_tot,
                savings, same_ramp_flag, ramp_pairs, shuttle_pairs, 'RRR',
                cur_raw=cur_row, pro_raw={},
            )
            (same_alts if same_ramp_flag else diff_alts).append(entry)

        same_alts.sort(key=lambda x:-x['savings']); diff_alts.sort(key=lambda x:-x['savings'])
        if same_alts:
            best=same_alts[0].copy(); best['all_same_alts']=same_alts; best['all_diff_alts']=diff_alts
            same_ramp_results.append(best)
        if diff_alts:
            best=diff_alts[0].copy(); best['all_same_alts']=same_alts; best['all_diff_alts']=diff_alts
            diff_ramp_results.append(best)

    same_ramp_results.sort(key=lambda x:-x['savings'])
    diff_ramp_results.sort(key=lambda x:-x['savings'])
    return (same_ramp_results, diff_ramp_results,
            dict(carrier_vol_total), dict(carrier_vol_ramp),
            ramp_pairs, shuttle_pairs)


# ═══════════════════════════════════════════════════════════════════════════════
# GEOGRAPHIC ALTERNATIVE FINDER
# ═══════════════════════════════════════════════════════════════════════════════
@st.cache_data
def build_geo_alternatives(_cf_df, _cfa_df):
    GEO_MAX_RADIUS = 200

    df_cols       = _cfa_df.columns.tolist()
    ramp_pairs    = get_ramp_col_pairs(df_cols)
    shuttle_pairs = get_shuttle_col_pairs(df_cols)

    cfa_index = defaultdict(list)
    for _, row in _cfa_df.iterrows():
        key = (row['Origin'].upper(), row['Fips'],
               normalize_col(row['Column']), row['Ramp'].upper())
        cfa_index[key].append(row.to_dict())

    # ── Build origin → carrier evidence from CFA ──────────────────────────────
    # If CFA has a row for Origin X at a ramp served by carrier Y,
    # then carrier Y can reach origin X. Use this to validate rate_needed rows.
    origin_carrier_evidence = defaultdict(set)
    for (orig, fips, col, ramp_mnem), rows in cfa_index.items():
        ramp_info = RAMP_COORDS.get(ramp_mnem[:5])
        if ramp_info:
            origin_carrier_evidence[orig].add(ramp_info['carrier'])

    # ── Also build origin → ramps seen in CFA (for any FIPS) ─────────────────
    # Used to check if alt ramp has ever appeared for this origin
    origin_ramp_evidence = defaultdict(set)
    for (orig, fips, col, ramp_mnem) in cfa_index.keys():
        origin_ramp_evidence[orig].add(ramp_mnem)

    def get_last_rail_carrier(cur_row):
        """Last active rail carrier in the flow — the one terminating at dest ramp."""
        active = {}
        for code, fc, _ in ramp_pairs:
            freight = clean_num(cur_row.get(fc, 0))
            if freight > 0:
                active[code] = freight
        if not active:
            return ''
        # Walk _RAIL_RATE_CARRIERS in reverse — last one with freight = terminal leg
        for carrier in reversed(_RAIL_RATE_CARRIERS):
            if carrier in active:
                return carrier
        return max(active, key=active.get)

    found_in_cfa  = []
    rate_needed   = []
    skipped_ramps = set()

    for _, cfrow in _cf_df.iterrows():
        orig    = cfrow['Origin'].upper()
        fips    = cfrow['Fips']
        col     = normalize_col(cfrow['Column'])
        ramp    = cfrow['Ramp'].upper()
        vol     = cfrow['_vol']
        county  = str(cfrow.get('County', '')).strip().title()
        st_code = str(cfrow.get('ST',     '')).strip()
        cf_flow = safe_str(cfrow.get('Flow', ''))
        if vol <= 0: continue

        # Strip suffix from ramp code for RAMP_COORDS lookup (e.g. 'PORTL' from 'PORTL-42')
        ramp_key = ramp[:5]
        cur_info = RAMP_COORDS.get(ramp_key)
        if cur_info is None:
            skipped_ramps.add(ramp)
            continue

        cfa_key     = (orig, fips, col, ramp)
        cur_matches = cfa_index.get(cfa_key, [])
        if not cur_matches:
            for k, v in cfa_index.items():
                if k[0] == orig and k[1] == fips and k[2] == col:
                    cur_matches = v; break
        if not cur_matches: continue

        cur_row = _pick_best_cfa_row(cur_matches, cf_flow)
        if cur_row is None: continue

        cur_cr        = calculate_cfa_cost(cur_row, vol, ramp_pairs, shuttle_pairs)
        cur_total     = cur_cr['total']
        if cur_total <= 0: continue

        cur_flow_str      = safe_str(cur_row.get('Flow', '')) or cf_flow
        cur_last_carrier  = get_last_rail_carrier(cur_row)
        cur_ramp_carrier  = cur_info['carrier']

        for alt_mnem, alt_info in RAMP_COORDS.items():
            if alt_mnem == ramp_key:                  continue
            if alt_info.get('status') == 'Inactive':  continue

            dist = haversine(cur_info['lat'], cur_info['lon'],
                             alt_info['lat'], alt_info['lon'])
            if dist > GEO_MAX_RADIUS: continue

            alt_carrier      = alt_info['carrier']
            cross_carrier    = cur_ramp_carrier != alt_carrier
            same_last_rail   = cur_last_carrier == alt_carrier

            # Railroad connectivity — does this carrier have evidence 
            # of serving this origin?
            carrier_evidence = alt_carrier in origin_carrier_evidence.get(orig, set())
            # Has this alt ramp ever appeared for this origin in CFA 
            # (any FIPS, any col)?
            ramp_seen_for_origin = alt_mnem in origin_ramp_evidence.get(orig, set())

            # Flow change description
            if not cross_carrier:
                flow_note = f"Same carrier ({alt_carrier}) — routing change only"
            else:
                flow_note = f"Flow will differ: {cur_ramp_carrier} → {alt_carrier}"

            base = {
                'origin'              : orig,
                'fips'                : fips,
                'county'              : county,
                'st'                  : st_code,
                'column'              : col,
                'volume'              : vol,
                'cur_ramp'            : ramp,
                'cur_carrier'         : cur_ramp_carrier,
                'cur_ramp_city'       : f"{cur_info['city']}, {cur_info['st']}",
                'cur_total'           : cur_total,
                'cur_cpu'             : cur_total / max(vol, 1),
                'cur_fd_carrier'      : cur_cr.get('fd_carrier', ''),
                'cur_flow'            : cur_flow_str,
                'cur_last_rail'       : cur_last_carrier,
                'alt_ramp'            : alt_mnem,
                'alt_carrier'         : alt_carrier,
                'alt_ramp_name'       : alt_info['name'],
                'alt_ramp_city'       : f"{alt_info['city']}, {alt_info['st']}",
                'ramp_dist_miles'     : round(dist, 1),
                'cross_carrier'       : cross_carrier,
                'same_last_rail'      : same_last_rail,
                'carrier_evidence'    : carrier_evidence,
                'ramp_seen_for_origin': ramp_seen_for_origin,
                'flow_note'           : flow_note,
            }

            # ── Exact CFA match — routing implicitly validated ────────────────
            alt_matches = cfa_index.get((orig, fips, col, alt_mnem), [])

            if alt_matches:
                alt_row   = _pick_best_cfa_row(alt_matches, cf_flow)
                if alt_row is None: continue
                alt_cr    = calculate_cfa_cost(alt_row, vol, ramp_pairs, shuttle_pairs)
                alt_total = alt_cr['total']
                if alt_total <= 0: continue
                savings   = cur_total - alt_total
                alt_last_rail = get_last_rail_carrier(alt_row)
                found_in_cfa.append({
                    **base,
                    'alt_total'      : alt_total,
                    'alt_cpu'        : alt_total / max(vol, 1),
                    'alt_fd_carrier' : alt_cr.get('fd_carrier', ''),
                    'alt_flow'       : safe_str(alt_row.get('Flow', '')),
                    'alt_last_rail'  : alt_last_rail,
                    'savings'        : savings,
                    'savings_pct'    : round((savings / cur_total * 100)
                                             if cur_total > 0 else 0.0, 1),
                })
            else:
                # ── Rate needed — only include if railroad can reach this origin ─
                # Must have either CFA evidence (carrier seen for this origin)
                # or the ramp itself has been seen for this origin in CFA
                if not carrier_evidence and not ramp_seen_for_origin:
                    continue  # No railroad evidence — skip, not a real opportunity
                rate_needed.append({**base})

    found_df  = pd.DataFrame(found_in_cfa) if found_in_cfa else pd.DataFrame()
    needed_df = pd.DataFrame(rate_needed)  if rate_needed  else pd.DataFrame()
    return found_df, needed_df, sorted(skipped_ramps)

def render_geo_tab(geo_found_df, geo_needed_df, skipped_ramps):
    st.markdown("### 🗺 Geographic Ramp Alternatives")
    st.caption(
        "Finds all active ramps within a configurable radius of each current ramp. "
        "Every comparison is controlled for **Origin + Column + FIPS** — same flow, different ramp. "
        "**Actionable** = CFA already has a rate at the nearby ramp.  "
        "**Rate Development** = nearby ramp exists but no CFA rate — "
        "get one quoted to create competition."
    )

    if geo_found_df.empty and geo_needed_df.empty:
        st.info("No geographic alternatives found — check that ramp codes in "
                "Country Flow match the MNEM codes in the ramp table.")
        if skipped_ramps:
            st.warning(f"Ramps not in coordinate table: {', '.join(skipped_ramps)}")
        return

    # ── Controls ──────────────────────────────────────────────────────────────
    c1, c2, c3 = st.columns(3)
    with c1:
        radius = st.slider("Search Radius (miles)", 10, 200, 75, 5, key='geo_radius')
    with c2:
        cross_only = st.checkbox(
            "Cross-carrier only", value=False, key='geo_cross',
            help="Only show alternatives that switch rail carrier"
        )
    with c3:
        min_savings = st.number_input(
            "Min CFA Savings $", min_value=0, value=0, step=1000, key='geo_minsav'
        )

    found_df  = (geo_found_df[ geo_found_df['ramp_dist_miles']  <= radius].copy()
                 if not geo_found_df.empty  else pd.DataFrame())
    needed_df = (geo_needed_df[geo_needed_df['ramp_dist_miles'] <= radius].copy()
                 if not geo_needed_df.empty else pd.DataFrame())

    if cross_only:
        if not found_df.empty:  found_df  = found_df[found_df['cross_carrier']]
        if not needed_df.empty: needed_df = needed_df[needed_df['cross_carrier']]

    if not found_df.empty:
        found_df = found_df[found_df['savings'] > 0].sort_values('savings', ascending=False)
    if not needed_df.empty:
        needed_df = needed_df.sort_values('volume', ascending=False)

    all_origins = sorted(set(
        (found_df['origin'].tolist()  if not found_df.empty  else []) +
        (needed_df['origin'].tolist() if not needed_df.empty else [])
    ))
    all_alt_carriers = sorted(set(
        (found_df['alt_carrier'].tolist()  if not found_df.empty  else []) +
        (needed_df['alt_carrier'].tolist() if not needed_df.empty else [])
    ))

    fc1, fc2 = st.columns(2)
    with fc1: sel_orig    = st.multiselect("Origin",      all_origins,      key='geo_orig')
    with fc2: sel_carrier = st.multiselect("Alt Carrier", all_alt_carriers, key='geo_alt_carrier')

    if sel_orig:
        if not found_df.empty:  found_df  = found_df[found_df['origin'].isin(sel_orig)]
        if not needed_df.empty: needed_df = needed_df[needed_df['origin'].isin(sel_orig)]
    if sel_carrier:
        if not found_df.empty:  found_df  = found_df[found_df['alt_carrier'].isin(sel_carrier)]
        if not needed_df.empty: needed_df = needed_df[needed_df['alt_carrier'].isin(sel_carrier)]
    if min_savings > 0 and not found_df.empty:
        found_df = found_df[found_df['savings'] >= min_savings]

    # ── Metrics ───────────────────────────────────────────────────────────────
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Actionable Switches",    f"{len(found_df):,}"               if not found_df.empty  else "0")
    m2.metric("CFA Savings Available",  fmt_c(found_df['savings'].sum())   if not found_df.empty  else "$—")
    m3.metric("Rate Dev Opportunities", f"{len(needed_df):,}"              if not needed_df.empty else "0")
    m4.metric("Volume Needing Rates",   f"{int(needed_df['volume'].sum()):,}" if not needed_df.empty else "0")

    if skipped_ramps:
        top  = ', '.join(sorted(skipped_ramps)[:10])
        more = f" …and {len(skipped_ramps)-10} more" if len(skipped_ramps) > 10 else ""
        st.info(f"ℹ️ {len(skipped_ramps)} ramp code(s) from Country Flow not in "
                f"coordinate table — skipped: **{top}{more}**")

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 1 — ACTIONABLE
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown("---")
    st.markdown("#### ✅ Actionable — Alternative Ramp Already Priced in CFA")
    st.caption(
        "CFA has a rate for this Origin + Column + FIPS at the nearby ramp. "
        "Savings are real and immediately actionable. "
        "**Ramp Dist** = straight-line distance between the two ramp locations."
    )

    if found_df.empty:
        st.info("No actionable geographic alternatives at this radius / filter combination.")
    else:
        disp = []
        for _, r in found_df.head(200).iterrows():
            disp.append({
                'Origin'         : r['origin'],
                'Col'            : r['column'],
                'County'         : r['county'],
                'ST'             : r['st'],
                'FIPS'           : r['fips'],
                'Volume'         : f"{int(r['volume']):,}",
                'Cur Ramp'       : r['cur_ramp'],
                'Cur Carrier'    : r['cur_carrier'],
                'Cur Location'   : r['cur_ramp_city'],
                'Cur FD'         : r['cur_fd_carrier'],
                'Cur CPU'        : f"${r['cur_cpu']:,.2f}",
                'Cur Total'      : fmt_c(r['cur_total']),
                'Cur Last Rail'  : r['cur_last_rail'],
                'Alt Rail'       : r['alt_last_rail'],
                'Flow Change'    : r['flow_note'],
                'Alt Ramp'       : r['alt_ramp'],
                'Alt Carrier'    : r['alt_carrier'],
                'Alt Location'   : r['alt_ramp_city'],
                'Alt FD'         : r['alt_fd_carrier'],
                'Alt CPU'        : f"${r['alt_cpu']:,.2f}",
                'Alt Total'      : fmt_c(r['alt_total']),
                'Ramp Dist (mi)' : f"{r['ramp_dist_miles']:.0f}",
                'Cross-Carrier'  : '✅' if r['cross_carrier'] else '',
                'Savings $'      : fmt_c(r['savings']),
                'Savings %'      : f"{r['savings_pct']:+.1f}%",
            })
        st.dataframe(pd.DataFrame(disp), use_container_width=True, hide_index=True)

        with st.expander("📊 Rollup by Current Ramp"):
            ramp_grp = found_df.groupby(['cur_ramp','cur_carrier']).agg(
                opportunities=('fips',    'count'),
                total_volume =('volume',  'sum'),
                total_savings=('savings', 'sum'),
            ).reset_index().sort_values('total_savings', ascending=False)

            best_idx = found_df.groupby('cur_ramp')['savings'].idxmax()
            best_alt = (found_df.loc[best_idx]
                        .set_index('cur_ramp')[['alt_ramp','alt_carrier']]
                        .to_dict('index'))

            rr_disp = []
            for _, r in ramp_grp.iterrows():
                ba = best_alt.get(r['cur_ramp'], {})
                rr_disp.append({
                    'Current Ramp'    : r['cur_ramp'],
                    'Carrier'         : r['cur_carrier'],
                    'Opportunities'   : int(r['opportunities']),
                    'Volume'          : f"{int(r['total_volume']):,}",
                    'Total Savings'   : fmt_c(r['total_savings']),
                    'Best Alt Ramp'   : ba.get('alt_ramp',    '—'),
                    'Best Alt Carrier': ba.get('alt_carrier', '—'),
                })
            st.dataframe(pd.DataFrame(rr_disp), use_container_width=True, hide_index=True)

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as writer:
            pd.DataFrame(disp).to_excel(writer, index=False, sheet_name='Geo Actionable')
        buf.seek(0)
        st.download_button(
            "⬇ Download Actionable Alternatives", data=buf,
            file_name="geo_ramp_alternatives.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_geo_found",
        )

    # ══════════════════════════════════════════════════════════════════════════
    # SECTION 2 — RATE DEVELOPMENT
    # ══════════════════════════════════════════════════════════════════════════
    st.markdown("---")
    st.markdown("#### 📋 Rate Development — Nearby Ramp Exists, No CFA Coverage")
    st.caption(
        "A geographically close ramp exists but Ford has no CFA rate for this lane at that ramp. "
        "Getting a rate quoted creates competition and could lower cost. "
        "Sorted by volume — largest flows = highest priority to develop. "
        "**Volume at Stake** = current annual spend — the leverage you take into the conversation."
    )

    if needed_df.empty:
        st.info("No rate development opportunities at this radius / filter combination.")
    else:
        nd_disp = []
        for _, r in needed_df.head(200).iterrows():
            nd_disp.append({
                'Origin'         : r['origin'],
                'Col'            : r['column'],
                'County'         : r['county'],
                'ST'             : r['st'],
                'FIPS'           : r['fips'],
                'Volume'         : f"{int(r['volume']):,}",
                'Cur Ramp'       : r['cur_ramp'],
                'Cur Carrier'    : r['cur_carrier'],
                'Cur Location'   : r['cur_ramp_city'],
                'Cur CPU'        : f"${r['cur_cpu']:,.2f}",
                'Cur Annual $'   : fmt_c(r['cur_total']),
                'Cur Last Rail'  : r['cur_last_rail'],
                'Flow Change'    : r['flow_note'],
                'Rail Evidence'  : '✅ Carrier seen for origin' if r['carrier_evidence']
                                else '📍 Ramp seen for origin',
                'Alt Ramp'       : r['alt_ramp'],
                'Alt Carrier'    : r['alt_carrier'],
                'Alt Location'   : r['alt_ramp_city'],
                'Ramp Dist (mi)' : f"{r['ramp_dist_miles']:.0f}",
                'Cross-Carrier'  : '✅' if r['cross_carrier'] else '',
                'Action'         : '📋 Get rate quoted',
            })
        st.dataframe(pd.DataFrame(nd_disp), use_container_width=True, hide_index=True)
    with st.expander("📊 Priority by Alt Carrier — Where to Focus Rate Development"):
        carrier_grp = needed_df.groupby('alt_carrier').agg(
            opportunities      =('fips',     'count'),
            total_volume       =('volume',   'sum'),
            distinct_origins   =('origin',   'nunique'),
            distinct_alt_ramps =('alt_ramp', 'nunique'),
        ).reset_index().sort_values('total_volume', ascending=False)

        cr_disp = []
        for _, r in carrier_grp.iterrows():
            cr_disp.append({
                'Alt Carrier'        : r['alt_carrier'],
                'Opportunities'      : int(r['opportunities']),
                'Volume at Stake'    : f"{int(r['total_volume']):,}",
                'Origins Affected'   : int(r['distinct_origins']),
                'Alt Ramps Available': int(r['distinct_alt_ramps']),
                'Action'             : 'Request CFA rate quotes',
            })
        st.dataframe(pd.DataFrame(cr_disp), use_container_width=True, hide_index=True)
        st.caption(
            "Volume at Stake = current volume flowing through other ramps that COULD "
            "route through this carrier's nearby ramp if Ford had a CFA rate. "
            "Use this number as leverage: 'Quote us a competitive rate and we can give you X units.'"
        )

        buf2 = io.BytesIO()
        with pd.ExcelWriter(buf2, engine='openpyxl') as writer:
            pd.DataFrame(nd_disp).to_excel(writer, index=False, sheet_name='Rate Development')
        buf2.seek(0)
        st.download_button(
            "⬇ Download Rate Development List", data=buf2,
            file_name="geo_rate_development.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_geo_needed",
        )


# ═══════════════════════════════════════════════════════════════════════════════
# FILTERS & TABLE BUILDERS
# ═══════════════════════════════════════════════════════════════════════════════
def apply_filters(rec_list, origins, states, min_sav, min_vol,
                  flow_kw, sources, show_flagged, limit=50):
    out = []; kw = flow_kw.strip().upper()
    for r in rec_list:
        if origins      and r['origin'] not in origins:  continue
        if states       and r['st']     not in states:   continue
        if sources      and r['source'] not in sources:  continue
        if r['savings'] < max(min_sav, _MIN_SAVINGS):    continue
        if r['volume']  < min_vol:                       continue
        if show_flagged and not r.get('flags'):          continue
        if kw and (kw not in r['current_flow'].upper() and
                   kw not in r['proposed_flow'].upper() and
                   kw not in r.get('what_changed','').upper()): continue
        out.append(r)
        if len(out) >= limit: break
    return out


def make_same_ramp_df(rec_list):
    rows = []
    for i, r in enumerate(rec_list, 1):
        warn = ''
        cr = r.get('current_breakdown', {})
        if cr.get('volume_adjusted'): warn += ' ⚠vol'
        if cr.get('cost_adjusted'):   warn += ' ⚠cost'
        if r.get('flags'):            warn += ' 🚩'
        cur_lbl = label_flow(r['current_flow'])
        pro_lbl = label_flow(r['proposed_flow'])
        flow_col = cur_lbl if cur_lbl == pro_lbl else f"{cur_lbl} → {pro_lbl}"
        rows.append({
            'Rank'         : i,
            'Source'       : r['source'],
            'Type'         : _type_label(r),
            'Origin'       : r['origin'],
            'RC'           : r.get('rc', ''),
            'County'       : r['county'],
            'ST'           : r['st'],
            'Col'          : r['column'],
            'Volume'       : int(r['volume']),
            'Ramp'         : r['current_ramp'],
            'Flow'         : flow_col,
            'What Changed' : r.get('what_changed', '—'),
            'Current $'    : fmt_c(r['current_cost']) + warn,
            'Proposed $'   : fmt_c(r['proposed_cost']),
            'Savings $'    : fmt_c(r['savings']),
            'Savings %'    : f"{r['savings_pct']:+.1f}%",
            'CPU Δ'        : f"${r['cpu_change']:,.2f}",
        })
    return pd.DataFrame(rows)


def make_diff_ramp_df(rec_list):
    rows = []
    for i, r in enumerate(rec_list, 1):
        warn = ''
        cr = r.get('current_breakdown', {})
        if cr.get('volume_adjusted'): warn += ' ⚠vol'
        if cr.get('cost_adjusted'):   warn += ' ⚠cost'
        if r.get('flags'):            warn += ' 🚩'
        rows.append({
            'Rank'          : i,
            'Source'        : r['source'],
            'Type'          : _type_label(r),
            'Origin'        : r['origin'],
            'RC'            : r.get('rc', ''),
            'County'        : r['county'],
            'ST'            : r['st'],
            'Col'           : r['column'],
            'Volume'        : int(r['volume']),
            'Current Flow'  : label_flow(r['current_flow']),
            'Proposed Flow' : label_flow(r['proposed_flow']),
            'Current $'     : fmt_c(r['current_cost']) + warn,
            'Proposed $'    : fmt_c(r['proposed_cost']),
            'Savings $'     : fmt_c(r['savings']),
            'Savings %'     : f"{r['savings_pct']:+.1f}%",
            'CPU Δ'         : f"${r['cpu_change']:,.2f}",
            'Ramp Change'   : f"{r['current_ramp']} → {r['proposed_ramp']}",
        })
    return pd.DataFrame(rows)


# ═══════════════════════════════════════════════════════════════════════════════
# DETAIL PANEL
# ═══════════════════════════════════════════════════════════════════════════════
def render_detail(rec, carrier_vol_total, carrier_vol_ramp):
    vol     = rec['volume']
    cur     = rec['current_breakdown']
    pro     = rec['proposed_breakdown']
    source  = rec.get('source', 'CFA')
    cur_raw = rec.get('current_raw', {})
    pro_raw = rec.get('proposed_raw', {})

    cur_cfa_vol = clean_num(cur_raw.get('Volume', 0)) if cur_raw else 0
    pro_cfa_vol = clean_num(pro_raw.get('Volume', 0)) if pro_raw else 0
    cur_scale   = (vol / cur_cfa_vol) if cur_cfa_vol > 0.000001 else 1.0
    pro_scale   = (vol / pro_cfa_vol) if pro_cfa_vol > 0.000001 else 1.0

    cur_rail_legs = get_rail_legs(cur_raw,   scale=cur_scale) if cur_raw else []
    pro_rail_legs = get_rail_legs(pro_raw,   scale=pro_scale) if pro_raw else []
    cur_sh_legs   = get_shuttle_legs(cur_raw, scale=cur_scale) if cur_raw else []
    pro_sh_legs   = get_shuttle_legs(pro_raw, scale=pro_scale) if pro_raw else []
    cur_fd        = get_fd_detail(cur_raw,   scale=cur_scale) if cur_raw else {}
    pro_fd        = get_fd_detail(pro_raw,   scale=pro_scale) if pro_raw else {}
    cur_id        = get_identity(cur_raw,    scale=cur_scale) if cur_raw else {}
    pro_id        = get_identity(pro_raw,    scale=pro_scale) if pro_raw else {}

    st.markdown("---")
    src_color = '#1e3a5f' if source == 'CFA' else '#2d5a27'
    st.markdown(
        f"### 📋 Flow Detail &nbsp;"
        f"<span style='background:{src_color};color:white;"
        f"padding:2px 8px;border-radius:4px;font-size:0.8em'>"
        f"{'📁 CFA' if source == 'CFA' else '📊 RRR'}</span>",
        unsafe_allow_html=True,
    )

    if rec.get('same_ramp') and rec.get('what_changed'):
        st.info(f"🔄 **What Changed:** {rec['what_changed']}")

    cur_is_dd = label_flow(rec['current_flow'])  == 'Dealer Direct'
    pro_is_dd = label_flow(rec['proposed_flow']) == 'Dealer Direct'
    if not rec.get('same_ramp'):
        if   cur_is_dd and not pro_is_dd:
            st.info("🚂 **Dealer Direct → Rail:** Adding a rail leg. Verify railcar availability.")
        elif not cur_is_dd and pro_is_dd:
            st.info("🚗 **Rail → Dealer Direct:** Removing the rail leg. Verify carhaul capacity.")

    if cur.get('volume_adjusted') or (pro_cfa_vol > 0 and abs(pro_cfa_vol - vol) > 0.01):
        st.warning(
            f"⚠️ **Volume scaling applied to leg detail:**  "
            f"Current CFA row = {int(cur_cfa_vol):,} units, "
            f"Proposed CFA row = {int(pro_cfa_vol):,} units, "
            f"Country Flow = {int(vol):,} units.  "
            f"All leg costs have been scaled to {int(vol):,} units.  "
            f"Rate/unit and Miles columns are per-unit and are NOT scaled."
        )
    if cur.get('cost_adjusted'):
        st.warning("⚠️ Carrier column sum deviated >1% from CFA summary — CFA total used.")
    for flag in rec.get('flags', []):
        st.warning(f"🚩 {flag}")

    # ── SECTION 1 — FLOW IDENTITY ─────────────────────────────────────────────
    with st.expander("🪪 Flow Identity", expanded=True):
        id1, id2, id3, id4 = st.columns(4)
        fips_disp = cur_id.get('fips', '') or rec.get('fips', '')
        rep_city  = cur_id.get('rep_city', '')
        rc_code   = rec.get('rc', '') or cur_id.get('rc', '')
        pt        = cur_id.get('pt', '')
        rrc       = cur_id.get('rrc', '')
        orig_rr   = cur_id.get('origin_rr', '')
        dest_rr   = cur_id.get('dest_rr', '')
        railcars  = cur_id.get('railcars', 0)
        rigs      = cur_id.get('carhaul_rigs', 0)
        clr       = cur_id.get('clr', 0)
        rlr       = cur_id.get('rlr', 0)
        t_rail_mi = cur_id.get('total_rail_miles', 0)
        t_miles   = cur_id.get('total_miles', 0)

        with id1:
            st.markdown(f"**Origin:** {rec['origin']}")
            st.markdown(f"**FIPS:** {fips_disp}")
            st.markdown(f"**County:** {rec['county']}, {rec['st']}")
            st.markdown(f"**Rep City:** {rep_city or '—'}")
        with id2:
            st.markdown(f"**RC:** {rc_code or '—'}")
            st.markdown(f"**Column:** {rec['column']}")
            st.markdown(f"**P/T:** {pt or '—'}  ·  **RRC:** {rrc or '—'}")
            st.markdown(f"**Volume:** {int(vol):,} units")
        with id3:
            st.markdown(f"**Origin RR:** {orig_rr or '—'}")
            st.markdown(f"**Dest RR:** {dest_rr or '—'}")
            st.markdown(f"**Railcars:** {int(railcars):,}" if railcars else "**Railcars:** —")
            st.markdown(f"**Carhaul Rigs:** {rigs:.1f}" if rigs else "**Carhaul Rigs:** —")
        with id4:
            st.markdown(f"**CLR:** {clr:.2f}" if clr else "**CLR:** —")
            st.markdown(f"**RLR:** {rlr:.2f}" if rlr else "**RLR:** —")
            st.markdown(f"**Total Rail Mi/unit:** {t_rail_mi:,.0f}" if t_rail_mi else "**Total Rail Mi/unit:** —")
            st.markdown(f"**Total Miles:** {t_miles:,.0f}" if t_miles else "**Total Miles:** —")

    # ── SECTION 2 — ROUTE DIAGRAM ─────────────────────────────────────────────
    with st.expander("🗺 Route", expanded=True):
        def build_route_parts(flow_str, ramp, fd_carrier, county, st_code, fips):
            tokens = [t for t in str(flow_str or '').split() if t.strip()]
            parts  = [f"**{t}**" for t in tokens]
            if ramp:       parts.append(f"📦 `{ramp}`")
            if fd_carrier: parts.append(f"**{fd_carrier}**")
            dest = f"📍 {county}, {st_code}"
            if fips:       dest += f" *(FIPS: {fips})*"
            parts.append(dest)
            return " ──► ".join(parts)

        cur_route = build_route_parts(
            rec['current_flow'],  rec['current_ramp'],
            rec['current_fd_carrier'],  rec['county'], rec['st'], fips_disp,
        )
        pro_route = build_route_parts(
            rec['proposed_flow'], rec['proposed_ramp'],
            rec['proposed_fd_carrier'], rec['county'], rec['st'], fips_disp,
        )
        rc1, rc2 = st.columns(2)
        with rc1:
            st.markdown("**Current Route**")
            st.markdown(cur_route)
        with rc2:
            st.markdown("**Proposed Route**")
            st.markdown(pro_route)

    # ── SECTION 3 — LEG-BY-LEG COST DETAIL ───────────────────────────────────
    with st.expander("💰 Leg-by-Leg Cost Detail", expanded=True):
        leg_tabs = st.tabs(["🚂 Rail Legs", "🔄 Shuttle", "🏭 Ramp Charges", "🚚 Final Delivery"])

        with leg_tabs[0]:
            if cur_rail_legs or pro_rail_legs:
                all_carriers = list(dict.fromkeys(
                    [l['carrier'] for l in cur_rail_legs] +
                    [l['carrier'] for l in pro_rail_legs]
                ))
                cur_ri = {l['carrier']: l for l in cur_rail_legs}
                pro_ri = {l['carrier']: l for l in pro_rail_legs}
                rail_rows = []
                for carrier in all_carriers:
                    cl = cur_ri.get(carrier, {})
                    pl = pro_ri.get(carrier, {})
                    cur_tot = cl.get('total', 0)
                    pro_tot = pl.get('total', 0)
                    rail_rows.append({
                        'Railroad'      : carrier,
                        'Cur Rate/unit' : fmt_r(cl.get('rate_pu', 0)),
                        'Cur Miles'     : f"{cl.get('miles',0):,}" if cl else '—',
                        'Cur Freight'   : fmt_c(cl.get('freight', 0)),
                        'Cur Fuel'      : fmt_c(cl.get('fuel', 0)),
                        'Cur Total'     : fmt_c(cur_tot),
                        'Pro Rate/unit' : fmt_r(pl.get('rate_pu', 0)),
                        'Pro Miles'     : f"{pl.get('miles',0):,}" if pl else '—',
                        'Pro Freight'   : fmt_c(pl.get('freight', 0)),
                        'Pro Fuel'      : fmt_c(pl.get('fuel', 0)),
                        'Pro Total'     : fmt_c(pro_tot),
                        'Δ'             : fmt_c(pro_tot - cur_tot),
                    })
                st.caption(f"All totals scaled to {int(vol):,} units. "
                           f"Rate/unit and Miles are per-unit values.")
                st.dataframe(pd.DataFrame(rail_rows),
                             use_container_width=True, hide_index=True)
            elif source == 'RRR':
                rail_bd = pro.get('rail_breakdown', {})
                if rail_bd:
                    carrier = list(rail_bd.keys())[0]
                    freight = rail_bd[carrier].get('freight', 0)
                    st.markdown(f"**RRR Rail Carrier:** {carrier}")
                    st.markdown(f"**Rate/unit:** {fmt_r(freight/max(vol,1))}  ·  "
                                f"**Total Freight:** {fmt_c(freight)}")
                else:
                    st.info("No rail leg data available for RRR proposals.")
            else:
                st.info("No rail legs found.")

        with leg_tabs[1]:
            if cur_sh_legs or pro_sh_legs:
                sh_rows = []
                for lg in cur_sh_legs:
                    sh_rows.append({
                        'Side'    : 'Current',
                        'Leg'     : f"Leg {lg['leg']}",
                        'Carrier' : lg['carrier'],
                        'Flow'    : lg.get('flow', ''),
                        'Miles'   : f"{lg['miles']:,}",
                        'Freight' : fmt_c(lg['freight']),
                        'Fuel'    : fmt_c(lg['fuel']),
                        'Total'   : fmt_c(lg['total']),
                    })
                for lg in pro_sh_legs:
                    sh_rows.append({
                        'Side'    : 'Proposed',
                        'Leg'     : f"Leg {lg['leg']}",
                        'Carrier' : lg['carrier'],
                        'Flow'    : lg.get('flow', ''),
                        'Miles'   : f"{lg['miles']:,}",
                        'Freight' : fmt_c(lg['freight']),
                        'Fuel'    : fmt_c(lg['fuel']),
                        'Total'   : fmt_c(lg['total']),
                    })
                st.caption(f"All totals scaled to {int(vol):,} units. Miles is per-unit.")
                st.dataframe(pd.DataFrame(sh_rows),
                             use_container_width=True, hide_index=True)
            else:
                st.info("No shuttle legs in this flow.")

        with leg_tabs[2]:
            cur_ramp_fr = cur_id.get('ramp_freight', 0)
            pro_ramp_fr = pro_id.get('ramp_freight', 0)
            ramp_rows = [
                {'Item': 'Ramp',          'Current': rec['current_ramp'],
                                           'Proposed': rec['proposed_ramp']},
                {'Item': 'Ramp Freight',  'Current': fmt_c(cur_ramp_fr),
                                           'Proposed': fmt_c(pro_ramp_fr)},
                {'Item': 'Ramp Frt/unit', 'Current': fmt_r(cur_ramp_fr / max(vol, 1)),
                                           'Proposed': fmt_r(pro_ramp_fr / max(vol, 1))},
                {'Item': 'Δ Total',       'Current': '',
                                           'Proposed': fmt_c(pro_ramp_fr - cur_ramp_fr)},
            ]
            if source == 'RRR':
                rrr_rc = pro.get('ramp_charges_pu', 0)
                ramp_rows.append({
                    'Item': 'RRR Ramp Charges/unit',
                    'Current': '—',
                    'Proposed': fmt_r(rrr_rc),
                })
            st.caption(f"Totals scaled to {int(vol):,} units.")
            st.dataframe(pd.DataFrame(ramp_rows),
                         use_container_width=True, hide_index=True)

        with leg_tabs[3]:
            if source == 'RRR':
                fd_rows = [
                    {'Item': 'FD Carrier',    'Current': cur.get('fd_carrier', '—'),
                                               'Proposed': pro.get('fd_carrier', '—')},
                    {'Item': 'Fixed/unit',    'Current': '—',
                                               'Proposed': fmt_r(pro.get('fd_fixed_pu', 0))},
                    {'Item': 'Var rate/mi',   'Current': '—',
                                               'Proposed': fmt_r(pro.get('fd_var_pu', 0))},
                    {'Item': 'Miles',         'Current': f"{cur.get('fd_miles',0):.0f}",
                                               'Proposed': f"{pro.get('fd_miles',0):.0f}"},
                    {'Item': 'FD Total/unit', 'Current': '—',
                                               'Proposed': fmt_r(
                                                   pro.get('fd_fixed_pu', 0) +
                                                   pro.get('fd_var_pu',   0) * pro.get('fd_miles', 0))},
                ]
            else:
                cfd = cur_fd; pfd = pro_fd
                cur_fd_total = cfd.get('freight', 0) + cfd.get('fuel', 0)
                pro_fd_total = pfd.get('freight', 0) + pfd.get('fuel', 0)
                fd_rows = [
                    {'Item': 'FD Carrier',  'Current': cfd.get('carrier', '—'),
                                             'Proposed': pfd.get('carrier', '—') or '—'},
                    {'Item': 'Fixed/unit',  'Current': fmt_r(cfd.get('fixed_pu', 0)),
                                             'Proposed': fmt_r(pfd.get('fixed_pu', 0))},
                    {'Item': 'Var rate/mi', 'Current': fmt_r(cfd.get('var_pu', 0)),
                                             'Proposed': fmt_r(pfd.get('var_pu', 0))},
                    {'Item': 'Miles',       'Current': f"{cfd.get('miles',0):.0f}",
                                             'Proposed': f"{pfd.get('miles',0):.0f}"},
                    {'Item': 'FD Freight',  'Current': fmt_c(cfd.get('freight', 0)),
                                             'Proposed': fmt_c(pfd.get('freight', 0))},
                    {'Item': 'FD Fuel',     'Current': fmt_c(cfd.get('fuel', 0)),
                                             'Proposed': fmt_c(pfd.get('fuel', 0))},
                    {'Item': 'FD Total',    'Current': fmt_c(cur_fd_total),
                                             'Proposed': fmt_c(pro_fd_total)},
                    {'Item': 'Δ FD Total',  'Current': '',
                                             'Proposed': fmt_c(pro_fd_total - cur_fd_total)},
                ]
            st.caption(f"Freight/Fuel totals scaled to {int(vol):,} units. "
                       f"Fixed/unit and Var/mi are per-unit rates.")
            st.dataframe(pd.DataFrame(fd_rows),
                         use_container_width=True, hide_index=True)

    # ── SECTION 4 — TOTAL COST SUMMARY ───────────────────────────────────────
    with st.expander("📊 Total Cost Summary", expanded=True):
        h1, h2, h3, h4 = st.columns(4)
        with h1: st.metric("Volume",       f"{int(vol):,} units")
        with h2: st.metric("Current CPU",  f"${rec['current_cpu']:,.2f}")
        with h3: st.metric("Proposed CPU", f"${rec['proposed_cpu']:,.2f}",
                            delta=f"${rec['cpu_change']:,.2f}")
        with h4: st.metric("Savings",      fmt_c(rec['savings']),
                            delta=f"{rec['savings_pct']:+.1f}%")

        summary_rows = [
            {'Item': 'Total Cost',
             'Current':  fmt_c(rec['current_cost']),
             'Proposed': fmt_c(rec['proposed_cost']),
             'Δ':        fmt_c(rec['proposed_cost'] - rec['current_cost'])},
            {'Item': '  Freight',
             'Current':  fmt_c(cur.get('total_freight', 0)),
             'Proposed': fmt_c(pro.get('total_freight', 0)),
             'Δ':        fmt_c(pro.get('total_freight', 0) - cur.get('total_freight', 0))},
            {'Item': '  Fuel',
             'Current':  fmt_c(cur.get('total_fuel', 0)),
             'Proposed': fmt_c(pro.get('total_fuel', 0)),
             'Δ':        fmt_c(pro.get('total_fuel', 0) - cur.get('total_fuel', 0))},
            {'Item': 'CPU',
             'Current':  f"${rec['current_cpu']:,.2f}",
             'Proposed': f"${rec['proposed_cpu']:,.2f}",
             'Δ':        f"${rec['cpu_change']:,.2f}"},
        ]
        if cur_id.get('cpu'):
            summary_rows.append({
                'Item': 'CFA CPU (raw)', 'Current': fmt_r(cur_id['cpu']),
                'Proposed': '—', 'Δ': '—',
            })
        st.dataframe(pd.DataFrame(summary_rows),
                     use_container_width=True, hide_index=True)

    # ── SECTION 5 — CARRIERS IMPACTED ────────────────────────────────────────
    with st.expander("🚛 Carriers Impacted"):
        def carriers_from_cr(cr):
            codes = set(cr.get('rail_breakdown', {}).keys())
            codes.update(cr.get('sh_breakdown', {}).keys())
            fd = cr.get('fd_carrier', '')
            if fd: codes.add(fd)
            return codes

        cur_carriers = carriers_from_cr(cur)
        pro_carriers = carriers_from_cr(pro)
        lost   = cur_carriers - pro_carriers
        gained = pro_carriers - cur_carriers
        stayed = cur_carriers & pro_carriers
        imp_rows = []
        for c in sorted(lost | gained | stayed):
            if   c in lost:   status, icon = 'LOST',   '🔴'
            elif c in gained: status, icon = 'GAINED', '🟢'
            else:             status, icon = 'STAYS',  '🟡'
            c_total   = carrier_vol_total.get(c, 0)
            pct_total = (vol / c_total * 100) if c_total > 0 else 0.0
            c_at_ramp = carrier_vol_ramp.get((c, rec['current_ramp']), 0)
            pct_ramp  = (vol / c_at_ramp * 100) if c_at_ramp > 0 else 0.0
            imp_rows.append({
                'Carrier'                     : f"{icon} {c}",
                'Status'                      : status,
                'Volume Impacted'             : f"{int(vol):,}",
                '% of Total Vol'              : f"{pct_total:.1f}%",
                f'% at {rec["current_ramp"]}' : f"{pct_ramp:.1f}%",
            })
        if imp_rows:
            st.dataframe(pd.DataFrame(imp_rows), use_container_width=True, hide_index=True)
        else:
            st.info("No carrier impact data found.")

    # ── SECTION 6 — MAP ───────────────────────────────────────────────────────
    with st.expander("🗺 Map"):
        o_lat = cur_id.get('o_lat', 0); o_lon = cur_id.get('o_lon', 0)
        d_lat = cur_id.get('d_lat', 0); d_lon = cur_id.get('d_lon', 0)
        if o_lat and d_lat:
            render_map(o_lat, o_lon, d_lat, d_lon,
                       f"{rec['origin']}",
                       f"{rec['county']}, {rec['st']} (FIPS {fips_disp})")
        else:
            st.info("Coordinates not available in CFA row.")

    # ── SECTION 7 — ALL ALTERNATIVES ─────────────────────────────────────────
    all_alts = sorted(
        rec.get('all_same_alts', []) + rec.get('all_diff_alts', []),
        key=lambda x: -x['savings']
    )
    if len(all_alts) > 1:
        with st.expander(f"🔁 All Alternatives ({len(all_alts)}) for {rec['county']}, {rec['st']}"):
            alt_disp = []
            for i, a in enumerate(all_alts, 1):
                flag_str = ' 🚩' if a.get('flags') else ''
                alt_disp.append({
                    'Rank'          : i,
                    'Source'        : a['source'],
                    'Type'          : _type_label(a),
                    'Change'        : '✅ Same Ramp' if a['same_ramp'] else '🔄 Ramp Change',
                    'Proposed Flow' : label_flow(a['proposed_flow']),
                    'Ramp'          : a['proposed_ramp'],
                    'What Changed'  : a.get('what_changed', '—') if a['same_ramp'] else '—',
                    'Proposed $'    : fmt_c(a['proposed_cost']),
                    'Savings $'     : fmt_c(a['savings']) + flag_str,
                    'Savings %'     : f"{a['savings_pct']:.1f}%",
                    'CPU'           : f"${a['proposed_cpu']:,.2f}",
                })
            st.dataframe(pd.DataFrame(alt_disp), use_container_width=True, hide_index=True)

    # ── SECTION 8 — EXPORT ───────────────────────────────────────────────────
    with st.expander("⬇ Export"):
        exp = []
        def er(*cols):
            row = list(cols)
            while len(row) < 4: row.append('')
            exp.append(row[:4])

        er('FORD CFA FLOW OPTIMIZATION — COMPARISON REPORT')
        er(f'Source: {source}')
        er(f'All leg costs scaled to {int(vol):,} units (Country Flow volume)')
        er('')
        er('IDENTITY')
        er('Origin',    rec['origin'],  'FIPS',      fips_disp)
        er('County',    rec['county'],  'State',     rec['st'])
        er('Rep City',  rep_city,       'RC',        rc_code)
        er('Column',    rec['column'],  'P/T',       pt)
        er('Volume',    int(vol),       'Origin RR', orig_rr)
        er('Dest RR',   dest_rr,        'Railcars',  int(railcars) if railcars else '—')
        er('CLR',       clr,            'RLR',       rlr)
        er('')
        er('',           'CURRENT',                        'PROPOSED', 'Δ')
        er('Flow',       label_flow(rec['current_flow']),
                         label_flow(rec['proposed_flow']),  '')
        er('Ramp',       rec['current_ramp'],               rec['proposed_ramp'], '')
        er('FD Carrier', rec['current_fd_carrier']  or '—', rec['proposed_fd_carrier'] or '—', '')
        er('Total Cost', fmt_c(rec['current_cost']),        fmt_c(rec['proposed_cost']),
                         fmt_c(rec['proposed_cost'] - rec['current_cost']))
        er('CPU',        f"${rec['current_cpu']:,.2f}",     f"${rec['proposed_cpu']:,.2f}",
                         f"${rec['cpu_change']:,.2f}")
        er('Savings $',  fmt_c(rec['savings']), '', '')
        er('Savings %',  f"{rec['savings_pct']:.1f}%", '', '')
        if rec.get('what_changed'):
            er('What Changed', rec['what_changed'], '', '')
        er('')
        er('RAIL LEGS (scaled)', 'Carrier', 'Current', 'Proposed')
        all_rail_carriers = list(dict.fromkeys(
            [l['carrier'] for l in cur_rail_legs] +
            [l['carrier'] for l in pro_rail_legs]
        ))
        cur_ri = {l['carrier']: l for l in cur_rail_legs}
        pro_ri = {l['carrier']: l for l in pro_rail_legs}
        for carrier in all_rail_carriers:
            cl = cur_ri.get(carrier, {}); pl = pro_ri.get(carrier, {})
            er(f"  {carrier} Rate/unit", fmt_r(cl.get('rate_pu', 0)),  fmt_r(pl.get('rate_pu', 0)), '')
            er(f"  {carrier} Miles",     cl.get('miles', '—'),          pl.get('miles', '—'), '')
            er(f"  {carrier} Freight",   fmt_c(cl.get('freight', 0)),   fmt_c(pl.get('freight', 0)),
                                         fmt_c(pl.get('freight', 0) - cl.get('freight', 0)))
            er(f"  {carrier} Fuel",      fmt_c(cl.get('fuel', 0)),      fmt_c(pl.get('fuel', 0)),
                                         fmt_c(pl.get('fuel', 0) - cl.get('fuel', 0)))
            er(f"  {carrier} Total",     fmt_c(cl.get('total', 0)),     fmt_c(pl.get('total', 0)),
                                         fmt_c(pl.get('total', 0) - cl.get('total', 0)))
        er('')
        er('SHUTTLE LEGS (scaled)', 'Carrier', 'Current', 'Proposed')
        for lg in cur_sh_legs:
            er(f"  Cur Leg {lg['leg']} {lg['carrier']}", '', fmt_c(lg['total']), '')
        for lg in pro_sh_legs:
            er(f"  Pro Leg {lg['leg']} {lg['carrier']}", '', '', fmt_c(lg['total']))
        er('')
        er('FINAL DELIVERY (scaled)', 'Current', 'Proposed', 'Δ')
        cfd = cur_fd; pfd = pro_fd
        er('  FD Carrier', cfd.get('carrier', '—'), pfd.get('carrier', '—'), '')
        er('  Fixed/unit', fmt_r(cfd.get('fixed_pu', 0)), fmt_r(pfd.get('fixed_pu', 0)), '')
        er('  Var/mi',     fmt_r(cfd.get('var_pu',   0)), fmt_r(pfd.get('var_pu',   0)), '')
        er('  Miles',      f"{cfd.get('miles',0):.0f}",   f"{pfd.get('miles',0):.0f}", '')
        cur_fd_t = cfd.get('freight', 0) + cfd.get('fuel', 0)
        pro_fd_t = pfd.get('freight', 0) + pfd.get('fuel', 0)
        er('  FD Total',   fmt_c(cur_fd_t), fmt_c(pro_fd_t), fmt_c(pro_fd_t - cur_fd_t))
        if rec.get('flags'):
            er(''); er('FLAGS')
            for f in rec['flags']: er(f)

        edf = pd.DataFrame(exp, columns=['A', 'B', 'C', 'D'])
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as writer:
            edf.to_excel(writer, index=False, header=False, sheet_name='Comparison')
        buf.seek(0)
        fname = (f"cfa_{rec['source'].lower()}_{rec['origin']}_"
                 f"{fips_disp}_{rec['county'].replace(' ','_')}.xlsx")
        st.download_button(
            label="⬇ Download as Excel", data=buf, file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"exp_{fips_disp}_{rec['origin']}_{rec['proposed_ramp']}_{source}",
        )


# ═══════════════════════════════════════════════════════════════════════════════
# TAB RENDERER
# ═══════════════════════════════════════════════════════════════════════════════
def render_tab(rec_list, tab_key, carrier_vol_total, carrier_vol_ramp, is_same_ramp):
    if not rec_list:
        st.info("No opportunities match the current filters."); return

    cfa_count  = sum(1 for r in rec_list if r['source'] == 'CFA')
    rrr_count  = sum(1 for r in rec_list if r['source'] == 'RRR')
    flag_count = sum(1 for r in rec_list if r.get('flags'))
    dd2rail    = sum(1 for r in rec_list if _type_label(r) == '🚂 DD→Rail')
    rail2dd    = sum(1 for r in rec_list if _type_label(r) == '🚗 Rail→DD')
    total_sav  = sum(r['savings'] for r in rec_list)
    total_vol  = sum(r['volume']  for r in rec_list)

    m1,m2,m3,m4,m5,m6,m7,m8 = st.columns(8)
    m1.metric("Opportunities", f"{len(rec_list)}")
    m2.metric("📁 CFA",        f"{cfa_count}")
    m3.metric("📊 RRR",        f"{rrr_count}")
    m4.metric("🚂 DD→Rail",    f"{dd2rail}")
    m5.metric("🚗 Rail→DD",    f"{rail2dd}")
    m6.metric("🚩 Flagged",    f"{flag_count}")
    m7.metric("Total Savings", fmt_c(total_sav))
    m8.metric("Volume",        f"{int(total_vol):,}")

    disp_df = make_same_ramp_df(rec_list) if is_same_ramp else make_diff_ramp_df(rec_list)
    event = st.dataframe(
        disp_df, use_container_width=True, hide_index=True,
        on_select="rerun", selection_mode="single-row",
        key=f"tbl_{tab_key}",
    )
    sel = []
    if hasattr(event, 'selection') and event.selection:
        sel = event.selection.rows or []

    if sel:
        idx = sel[0]
        if 0 <= idx < len(rec_list):
            render_detail(rec_list[idx], carrier_vol_total, carrier_vol_ramp)
    else:
        st.caption(
            "👆 Click any row to expand the full flow card — legs, rates, map, carriers.  "
            "⚠vol = volume scaled · ⚠cost = cost adjusted · 🚩 = data flag"
        )

def render_whatif_tab(cf_df, cfa_df, ramp_pairs, shuttle_pairs):
    st.markdown("### 🔀 What-If Scenario Builder")
    st.caption(
        "Drill down from origin → destination → volume → where to move it. "
        "Each step loads from your actual Country Flow and CFA data."
    )

    df_cols = cfa_df.columns.tolist()
    if not ramp_pairs:
        ramp_pairs    = get_ramp_col_pairs(df_cols)
        shuttle_pairs = get_shuttle_col_pairs(df_cols)

    # Build CFA index once
    cfa_index = defaultdict(list)
    for _, row in cfa_df.iterrows():
        key = (row['Origin'].upper(), row['Fips'],
               normalize_col(row['Column']), row['Ramp'].upper())
        cfa_index[key].append(row.to_dict())

    # ── STEP 1: Pick Origin ───────────────────────────────────────────────────
    st.markdown("#### Step 1 — Pick Origin")
    all_origins = ['— Select —'] + sorted(cf_df['Origin'].str.upper().unique())
    sel_origin  = st.selectbox("Origin Plant / Source", all_origins, key='wi_origin')

    if sel_origin == '— Select —':
        st.info("Select an origin to load its shipping lanes.")
        return

    origin_cf = cf_df[cf_df['Origin'].str.upper() == sel_origin].copy()

    # Quick stats for selected origin
    oa, ob, oc, od = st.columns(4)
    oa.metric("Total Volume",  f"{int(origin_cf['_vol'].sum()):,}")
    ob.metric("Ramps Served",  f"{origin_cf['Ramp'].nunique()}")
    oc.metric("States",        f"{origin_cf['ST'].nunique()}")
    od.metric("Lanes (FIPS)",  f"{origin_cf['Fips'].nunique()}")

    # ── STEP 2: Filter what to move ──────────────────────────────────────────
    st.markdown("#### Step 2 — Select What to Move")
    st.caption("Narrow down by state, ramp, or county. Leave blank to include everything.")

    f1, f2, f3 = st.columns(3)
    with f1:
        all_states   = sorted(origin_cf['ST'].unique())
        sel_states   = st.multiselect("State(s)", all_states, key='wi_states')
    with f2:
        all_cur_ramps = sorted(origin_cf['Ramp'].str.upper().unique())
        sel_cur_ramps = st.multiselect("Current Ramp(s)", all_cur_ramps, key='wi_cur_ramps')
    with f3:
        all_counties  = sorted(origin_cf['County'].str.strip().str.title().unique())
        sel_counties  = st.multiselect("County(s)", all_counties, key='wi_counties')

    # Apply filters live — no button needed here
    vol_mask = pd.Series(True, index=origin_cf.index)
    if sel_states:    vol_mask &= origin_cf['ST'].isin(sel_states)
    if sel_cur_ramps: vol_mask &= origin_cf['Ramp'].str.upper().isin(sel_cur_ramps)
    if sel_counties:  vol_mask &= origin_cf['County'].str.strip().str.title().isin(sel_counties)

    sel_vol_df = origin_cf[vol_mask].copy()

    if sel_vol_df.empty:
        st.warning("No volume matches the current filters.")
        return

    total_sel_vol = sel_vol_df['_vol'].sum()

    # Show volume breakdown by ramp — live, before running
    with st.expander("📋 Volume in Scope (by Ramp)", expanded=True):
        ramp_summ = (
            sel_vol_df.groupby(['Ramp', 'ST'])
            .agg(counties=('County', 'nunique'),
                 lanes   =('Fips',   'count'),
                 volume  =('_vol',   'sum'))
            .reset_index()
            .sort_values('volume', ascending=False)
        )
        ramp_disp = []
        for _, r in ramp_summ.iterrows():
            # Current cost at this ramp (sum over matching CFA rows)
            ramp_rows_cf = sel_vol_df[sel_vol_df['Ramp'].str.upper() == r['Ramp'].upper()]
            cur_cost_ramp = 0.0
            for _, cfrow in ramp_rows_cf.iterrows():
                orig = cfrow['Origin'].upper(); fips = cfrow['Fips']
                col  = normalize_col(cfrow['Column']); ramp = cfrow['Ramp'].upper()
                vol  = cfrow['_vol']
                cf_flow = safe_str(cfrow.get('Flow', ''))
                matches = cfa_index.get((orig, fips, col, ramp), [])
                if not matches:
                    for k, v in cfa_index.items():
                        if k[0] == orig and k[1] == fips and k[2] == col:
                            matches = v; break
                if not matches: continue
                row_match = _pick_best_cfa_row(matches, cf_flow)
                if row_match is None: continue
                cr = calculate_cfa_cost(row_match, vol, ramp_pairs, shuttle_pairs)
                cur_cost_ramp += cr['total']
            ramp_disp.append({
                'Ramp'         : r['Ramp'],
                'State'        : r['ST'],
                'Counties'     : int(r['counties']),
                'Lanes'        : int(r['lanes']),
                'Volume'       : f"{int(r['volume']):,}",
                'Vol %'        : f"{r['volume'] / total_sel_vol * 100:.1f}%",
                'Current Cost' : fmt_c(cur_cost_ramp),
                'CPU'          : fmt_c(cur_cost_ramp / max(r['volume'], 1)),
            })
        st.dataframe(pd.DataFrame(ramp_disp), use_container_width=True, hide_index=True)

    # ── STEP 3: How much to move ──────────────────────────────────────────────
    st.markdown("#### Step 3 — How Much to Move")
    vol_pct = st.slider(
        "% of selected volume to move", 
        min_value=10, max_value=100, value=100, step=10,
        key='wi_vol_pct',
        help="100% = move all selected volume. 50% = move half, scaled proportionally across all lanes."
    )
    scale_factor = vol_pct / 100.0
    moved_vol    = total_sel_vol * scale_factor
    st.caption(f"Moving **{int(moved_vol):,}** of **{int(total_sel_vol):,}** selected units ({vol_pct}%)")

    # ── STEP 4: Where to move it ──────────────────────────────────────────────
    st.markdown("#### Step 4 — Move To")
    cur_ramps_in_scope = set(sel_vol_df['Ramp'].str.upper().unique())
    all_ramps_cfa      = sorted(cfa_df['Ramp'].str.upper().unique())
    to_ramp_opts       = [r for r in all_ramps_cfa if r not in cur_ramps_in_scope]

    t1, t2 = st.columns(2)
    with t1:
        to_ramp = st.selectbox(
            "Destination Ramp",
            ['— Find Best Available —'] + to_ramp_opts,
            key='wi_to_ramp',
            help="Pick a specific ramp, or let the tool find the cheapest available CFA rate."
        )
    with t2:
        min_savings_filter = st.number_input(
            "Min Savings $ to show", min_value=0, value=0, step=500, key='wi_min_sav'
        )

    # ── RUN ───────────────────────────────────────────────────────────────────
    if st.button("▶ Run Scenario", key='wi_run', type='primary'):

        to_ramps_check = (
            [to_ramp] if to_ramp != '— Find Best Available —'
            else to_ramp_opts
        )

        results  = []
        no_rate  = []

        for _, cfrow in sel_vol_df.iterrows():
            orig    = cfrow['Origin'].upper()
            fips    = cfrow['Fips']
            col     = normalize_col(cfrow['Column'])
            ramp    = cfrow['Ramp'].upper()
            vol     = cfrow['_vol'] * scale_factor
            county  = str(cfrow.get('County', '')).strip().title()
            st_code = str(cfrow.get('ST', '')).strip()
            cf_flow = safe_str(cfrow.get('Flow', ''))
            rc_code = safe_str(cfrow.get('RC', ''))
            if vol <= 0: continue

            # Current cost
            cur_matches = cfa_index.get((orig, fips, col, ramp), [])
            if not cur_matches:
                for k, v in cfa_index.items():
                    if k[0] == orig and k[1] == fips and k[2] == col:
                        cur_matches = v; break
            if not cur_matches: continue

            cur_row = _pick_best_cfa_row(cur_matches, cf_flow)
            if cur_row is None: continue
            cur_cr    = calculate_cfa_cost(cur_row, vol, ramp_pairs, shuttle_pairs)
            cur_total = cur_cr['total']
            if cur_total <= 0: continue

            # Find best TO ramp
            best_alt = None
            for tr in to_ramps_check:
                if tr == ramp: continue
                alt_matches = cfa_index.get((orig, fips, col, tr), [])
                if not alt_matches: continue
                alt_row = _pick_best_cfa_row(alt_matches, cf_flow)
                if alt_row is None: continue
                alt_cr    = calculate_cfa_cost(alt_row, vol, ramp_pairs, shuttle_pairs)
                alt_total = alt_cr['total']
                if alt_total <= 0: continue
                savings = cur_total - alt_total
                if best_alt is None or savings > best_alt['savings']:
                    best_alt = {
                        'alt_ramp'    : tr,
                        'alt_carrier' : alt_cr.get('fd_carrier', ''),
                        'alt_total'   : alt_total,
                        'alt_cpu'     : alt_total / max(vol, 1),
                        'alt_fd'      : alt_cr.get('fd_carrier', ''),
                        'alt_flow'    : safe_str(alt_row.get('Flow', '')),
                        'savings'     : savings,
                        'savings_pct' : round((savings / cur_total * 100)
                                              if cur_total > 0 else 0.0, 1),
                    }

            if best_alt:
                if best_alt['savings'] >= min_savings_filter:
                    results.append({
                        'origin'      : orig,
                        'fips'        : fips,
                        'county'      : county,
                        'st'          : st_code,
                        'column'      : col,
                        'rc'          : rc_code,
                        'volume'      : vol,
                        'cur_ramp'    : ramp,
                        'cur_carrier' : cur_cr.get('fd_carrier', ''),
                        'cur_fd'      : cur_cr.get('fd_carrier', ''),
                        'cur_total'   : cur_total,
                        'cur_cpu'     : cur_total / max(vol, 1),
                        **best_alt,
                    })
            else:
                no_rate.append({
                    'origin'   : orig,
                    'fips'     : fips,
                    'county'   : county,
                    'st'       : st_code,
                    'column'   : col,
                    'volume'   : vol,
                    'cur_ramp' : ramp,
                    'note'     : 'No CFA rate found at any TO ramp for this lane',
                })

        # ── Results ───────────────────────────────────────────────────────────
        st.markdown("---")
        st.markdown("#### Results")

        if not results and not no_rate:
            st.info("No matches found. The selected TO ramp may not have CFA rates for these lanes.")
            return

        res_df         = pd.DataFrame(results) if results else pd.DataFrame()
        total_cur_cost = res_df['cur_total'].sum()  if not res_df.empty else 0
        total_alt_cost = res_df['alt_total'].sum()  if not res_df.empty else 0
        total_savings  = res_df['savings'].sum()    if not res_df.empty else 0
        total_vol_res  = res_df['volume'].sum()     if not res_df.empty else 0
        no_rate_vol    = sum(r['volume'] for r in no_rate)

        m1, m2, m3, m4, m5, m6 = st.columns(6)
        m1.metric("Lanes",          f"{len(results):,}")
        m2.metric("Volume Moved",   f"{int(total_vol_res):,}")
        m3.metric("Current Cost",   fmt_c(total_cur_cost))
        m4.metric("Proposed Cost",  fmt_c(total_alt_cost))
        m5.metric("Total Savings",  fmt_c(total_savings),
                  delta=f"{(total_savings / total_cur_cost * 100):+.1f}%"
                  if total_cur_cost > 0 else None)
        m6.metric("No Rate Lanes",  f"{len(no_rate):,}",
                  delta=f"{int(no_rate_vol):,} units" if no_rate else None,
                  delta_color="inverse")

        if not res_df.empty:
            # Detail table
            disp = []
            for _, r in res_df.sort_values('savings', ascending=False).iterrows():
                disp.append({
                    'Origin'       : r['origin'],
                    'Col'          : r['column'],
                    'County'       : r['county'],
                    'ST'           : r['st'],
                    'FIPS'         : r['fips'],
                    'RC'           : r['rc'],
                    'Volume'       : f"{int(r['volume']):,}",
                    'From Ramp'    : r['cur_ramp'],
                    'From FD'      : r['cur_fd'],
                    'From CPU'     : f"${r['cur_cpu']:,.2f}",
                    'From Total'   : fmt_c(r['cur_total']),
                    'To Ramp'      : r['alt_ramp'],
                    'To FD'        : r['alt_fd'],
                    'To CPU'       : f"${r['alt_cpu']:,.2f}",
                    'To Total'     : fmt_c(r['alt_total']),
                    'Savings $'    : fmt_c(r['savings']),
                    'Savings %'    : f"{r['savings_pct']:+.1f}%",
                })

            disp_df = pd.DataFrame(disp)
            event = st.dataframe(
                disp_df, use_container_width=True, hide_index=True,
                on_select="rerun", selection_mode="single-row",
                key="wi_res_tbl",
            )
            sel = []
            if hasattr(event, 'selection') and event.selection:
                sel = event.selection.rows or []

            if sel:
                idx = sel[0]
                sorted_res_df = res_df.sort_values('savings', ascending=False).reset_index(drop=True)
                if 0 <= idx < len(sorted_res_df):
                    row = sorted_res_df.iloc[idx]
                    st.markdown("---")
                    st.markdown(
                        f"#### 🔍 Flow Detail — {row['county']}, {row['st']}  |  "
                        f"FIPS {row['fips']}  |  Col {row['column']}  |  "
                        f"Vol {int(row['volume']):,}"
                    )
                    left, right = st.columns(2)

                    with left:
                        st.markdown(f"##### 📍 Current — {row['cur_ramp']}")
                        st.markdown(f"**FD Carrier:** {row['cur_fd'] or '—'}")
                        st.markdown(f"**CPU:** ${row['cur_cpu']:,.2f}")
                        st.markdown(f"**Total Cost:** {fmt_c(row['cur_total'])}")

                    with right:
                        st.markdown(f"##### ✅ Proposed — {row['alt_ramp']}")
                        st.markdown(f"**FD Carrier:** {row['alt_fd'] or '—'}")
                        st.markdown(f"**Flow:** {row['alt_flow'] or '—'}")
                        st.markdown(f"**CPU:** ${row['alt_cpu']:,.2f}")
                        st.markdown(f"**Total Cost:** {fmt_c(row['alt_total'])}")

                    st.markdown(
                        f"**Savings:** {fmt_c(row['savings'])} ({row['savings_pct']:+.1f}%)  |  "
                        f"**Volume Moved:** {int(row['volume']):,}  |  "
                        f"**RC:** {row['rc'] or '—'}"
                    )
            else:
                st.caption("👆 Click any row to expand the full current vs proposed flow detail.")

            st.dataframe(pd.DataFrame(disp), use_container_width=True, hide_index=True)

            # Rollup by FROM ramp
            with st.expander("📊 Rollup by Current Ramp"):
                grp = (res_df.groupby('cur_ramp')
                       .agg(lanes     =('fips',      'count'),
                            volume    =('volume',    'sum'),
                            cur_cost  =('cur_total', 'sum'),
                            alt_cost  =('alt_total', 'sum'),
                            savings   =('savings',   'sum'))
                       .reset_index()
                       .sort_values('savings', ascending=False))
                grp_disp = []
                for _, r in grp.iterrows():
                    grp_disp.append({
                        'From Ramp'  : r['cur_ramp'],
                        'Lanes'      : int(r['lanes']),
                        'Volume'     : f"{int(r['volume']):,}",
                        'Current $'  : fmt_c(r['cur_cost']),
                        'Proposed $' : fmt_c(r['alt_cost']),
                        'Savings $'  : fmt_c(r['savings']),
                        'Savings %'  : f"{(r['savings'] / r['cur_cost'] * 100):+.1f}%"
                                       if r['cur_cost'] > 0 else '—',
                    })
                st.dataframe(pd.DataFrame(grp_disp), use_container_width=True, hide_index=True)

            # Rollup by TO ramp
            with st.expander("📊 Rollup by Destination Ramp"):
                grp2 = (res_df.groupby('alt_ramp')
                        .agg(lanes   =('fips',      'count'),
                             volume  =('volume',    'sum'),
                             savings =('savings',   'sum'),
                             alt_cost=('alt_total', 'sum'))
                                                     .reset_index()
                        .sort_values('savings', ascending=False))
                grp2_disp = []
                for _, r in grp2.iterrows():
                    grp2_disp.append({
                        'To Ramp'    : r['alt_ramp'],
                        'Lanes'      : int(r['lanes']),
                        'Volume'     : f"{int(r['volume']):,}",
                        'Proposed $' : fmt_c(r['alt_cost']),
                        'Savings $'  : fmt_c(r['savings']),
                    })
                st.dataframe(pd.DataFrame(grp2_disp), use_container_width=True, hide_index=True)

            # Export
            out = io.BytesIO()
            pd.DataFrame(disp).to_excel(out, index=False)
            st.download_button(
                "⬇️ Download Scenario Results",
                data=out.getvalue(),
                file_name="whatif_scenario.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        if no_rate:
            with st.expander(f"🔴 {len(no_rate)} Lane(s) With No Rate at TO Ramp"):
                st.dataframe(pd.DataFrame(no_rate), use_container_width=True, hide_index=True)

# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    st.title("🚗 Ford CFA Flow Optimization Tool")
    ...

    # ADD THIS BLOCK
    with st.sidebar:
        st.markdown("### 📂 Upload Files")
        cfa_file = st.file_uploader("CFA File (.xlsm / .xlsx)", type=['xlsx','xlsm'], key='cfa_upload')
        rrr_file = st.file_uploader("RRR File (.xlsx)",          type=['xlsx'],       key='rrr_upload')
        cf_file  = st.file_uploader("Country Flow File (.xlsx)", type=['xlsx'],       key='cf_upload')
        if not all([cfa_file, rrr_file, cf_file]):
            st.info("⬆️ Upload all three files to begin.")
            st.stop()


    errors = []
    with st.spinner("Loading files…"):
        try:    cf_df    = load_country_flow(cf_file.read())
        except Exception as e: errors.append(f"Country Flow: {e}"); cf_df = pd.DataFrame()
        try:    cfa_df   = load_cfa(cfa_file.read())
        except Exception as e: errors.append(f"CFA: {e}"); cfa_df = pd.DataFrame()
        try:    rrr_data = load_rrr(rrr_file.read())
        except Exception as e:
            errors.append(f"RRR: {e}")
            rrr_data = {k: pd.DataFrame() for k in ('r2r','r2t','r2c','via','rc')}

    for err in errors: st.error(f"❌ {err}")
    if cf_df.empty or cfa_df.empty: return

    st.sidebar.success(
        f"✅ Country Flow: {len(cf_df):,} rows\n\n"
        f"✅ CFA: {len(cfa_df):,} rows\n\n"
        f"✅ RRR R2R: {len(rrr_data['r2r']):,} rows\n\n"
        f"✅ RRR R2T: {len(rrr_data['r2t']):,} rows\n\n"
        f"✅ Ramp Charges: {len(rrr_data['rc']):,} rows"
    )

    with st.spinner("Computing savings opportunities…"):
        try:
            (same_all, diff_all,
             c_vol_total, c_vol_ramp,
             ramp_pairs, shuttle_pairs) = build_comparisons(
                cf_df, cfa_df,
                rrr_data['r2r'], rrr_data['r2t'], rrr_data['rc']
            )
        except Exception as e:
            st.error(f"Computation error: {e}"); st.exception(e); return

    with st.spinner("Finding geographic alternatives…"):
        try:
            geo_found_df, geo_needed_df, skipped_ramps = build_geo_alternatives(
                cf_df, cfa_df
            )
        except Exception as e:
            st.warning(f"Geographic analysis unavailable: {e}")
            geo_found_df = geo_needed_df = pd.DataFrame()
            skipped_ramps = []

    if not same_all and not diff_all:
        st.warning("No savings opportunities found.")
        with st.expander("🔍 Diagnostics"):
            st.write(f"**Country Flow rows:** {len(cf_df):,}")
            st.write(f"**CFA rows:** {len(cfa_df):,}")
            st.write("**Country Flow Origins:**", cf_df['Origin'].unique().tolist())
            st.write("**CFA Origins:**",          cfa_df['Origin'].unique().tolist())
            st.write("**Sample CF Fips:**",        cf_df['Fips'].head(5).tolist())
            st.write("**Sample CFA Fips:**",       cfa_df['Fips'].head(5).tolist())
        return

    with st.sidebar:
        st.header("🔧 Filters")
        all_origins = sorted(set(r['origin'] for r in same_all + diff_all))
        all_states  = sorted(set(r['st']     for r in same_all + diff_all))
        sel_origins = st.multiselect("Origin", all_origins, default=all_origins)
        sel_states  = st.multiselect("State",  all_states,  default=all_states)
        sel_sources = st.multiselect("Source", ['CFA','RRR'], default=['CFA','RRR'])
        sel_types   = st.multiselect(
            "Transition Type",
            ['🚂 DD→Rail','🚗 Rail→DD','🔄 Rail→Rail'],
            default=['🚂 DD→Rail','🚗 Rail→DD','🔄 Rail→Rail'],
        )
        min_sav  = st.number_input("Min Savings $",      min_value=0, value=0,
                                    step=5_000)
        min_vol  = st.number_input("Min Volume (units)",  min_value=1,
                                    value=_MIN_VOLUME_DEFAULT, step=10)
        flow_kw  = st.text_input("Flow / Carrier Keyword",
                                  placeholder="e.g. URSI, NS, NWBST")
        show_flagged = st.checkbox("Show flagged rows only 🚩", value=False)
        st.markdown("---")
        grand = sum(r['savings'] for r in same_all[:50] + diff_all[:50])
        st.markdown(f"Same-ramp opps:   **{len(same_all)}**")
        st.markdown(f"Ramp-change opps: **{len(diff_all)}**")
        st.markdown(f"Top-50 combined savings: **{fmt_c(grand)}**")

    same_filt = apply_filters(same_all, sel_origins, sel_states, min_sav, min_vol,
                               flow_kw, sel_sources, show_flagged)
    diff_filt = apply_filters(diff_all, sel_origins, sel_states, min_sav, min_vol,
                               flow_kw, sel_sources, show_flagged)
    if sel_types and len(sel_types) < 3:
        same_filt = [r for r in same_filt if _type_label(r) in sel_types]
        diff_filt = [r for r in diff_filt if _type_label(r) in sel_types]

    tab1, tab2, tab3, tab4 = st.tabs([
        f"✅  Same Ramp  ({len(same_filt)} opportunities)",
        f"🔄  Ramp Change  ({len(diff_filt)} opportunities)",
        "🗺  Geographic Alternatives","🔀 What-If"
])
    with tab1:
        st.markdown("Ramp unchanged — **What Changed** column shows exactly what differs.")
        render_tab(same_filt, 'same', c_vol_total, c_vol_ramp, is_same_ramp=True)
    with tab2:
        st.markdown("Vehicles unload at a **different destination ramp**.")
        render_tab(diff_filt, 'diff', c_vol_total, c_vol_ramp, is_same_ramp=False)
    with tab3:
        render_geo_tab(geo_found_df, geo_needed_df, skipped_ramps)
    with tab4:
        render_whatif_tab(cf_df, cfa_df, ramp_pairs, shuttle_pairs)

if __name__ == '__main__':
    main()

